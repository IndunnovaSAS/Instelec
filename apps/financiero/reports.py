"""Generador de cuadro de costos para facturación."""

import io
from datetime import date
from decimal import Decimal
from typing import Any

from django.db.models import Sum, Count, F
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter

from apps.financiero.models import CostoRecurso, Presupuesto, EjecucionCosto, CicloFacturacion
from apps.actividades.models import Actividad
from apps.cuadrillas.models import Cuadrilla, CuadrillaMiembro, Vehiculo
from apps.lineas.models import Linea


class CuadroCostosGenerator:
    """Genera cuadro de costos para facturación mensual."""

    def __init__(self, anio: int, mes: int, linea_id: str = None):
        self.anio = anio
        self.mes = mes
        self.linea_id = linea_id
        self.linea = Linea.objects.get(id=linea_id) if linea_id else None
        self.data = {}

    def consolidar_datos(self) -> dict[str, Any]:
        """Consolida todos los costos del período."""
        # Filtrar actividades del período
        actividades_qs = Actividad.objects.filter(
            fecha_programada__year=self.anio,
            fecha_programada__month=self.mes,
            estado='COMPLETADA'
        )

        if self.linea:
            actividades_qs = actividades_qs.filter(linea=self.linea)

        actividades = actividades_qs.select_related(
            'linea', 'torre', 'tipo_actividad', 'cuadrilla'
        )

        # Calcular costos
        self.data = {
            'periodo': f"{self._get_mes_nombre(self.mes)} {self.anio}",
            'linea': self.linea,
            'fecha_generacion': date.today(),

            # Resumen
            'resumen': {
                'actividades_completadas': actividades.count(),
                'torres_intervenidas': actividades.values('torre').distinct().count(),
                'dias_trabajados': self._calcular_dias_trabajados(actividades),
            },

            # Costos de personal
            'personal': self._calcular_costos_personal(actividades),

            # Costos de vehículos
            'vehiculos': self._calcular_costos_vehiculos(actividades),

            # Costos de materiales
            'materiales': self._calcular_costos_materiales(actividades),

            # Costos de herramientas
            'herramientas': self._calcular_costos_herramientas(actividades),

            # Otros costos
            'otros': self._calcular_otros_costos(actividades),

            # Totales
            'totales': {},
        }

        # Calcular totales
        self._calcular_totales()

        return self.data

    def generar_excel(self) -> bytes:
        """Genera el cuadro de costos en formato Excel."""
        if not self.data:
            self.consolidar_datos()

        wb = Workbook()
        self._configurar_estilos(wb)

        # Hoja de resumen
        self._crear_hoja_resumen(wb)

        # Hoja de personal
        self._crear_hoja_personal(wb)

        # Hoja de vehículos
        self._crear_hoja_vehiculos(wb)

        # Hoja de materiales
        self._crear_hoja_materiales(wb)

        # Hoja de detalle por actividad
        self._crear_hoja_detalle(wb)

        # Guardar
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer.getvalue()

    def _configurar_estilos(self, wb: Workbook):
        """Configura estilos globales del workbook."""
        # Estilo de encabezado
        header_style = NamedStyle(name='header_style')
        header_style.font = Font(bold=True, color='FFFFFF', size=11)
        header_style.fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
        header_style.alignment = Alignment(horizontal='center', vertical='center')
        header_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        wb.add_named_style(header_style)

        # Estilo de moneda
        currency_style = NamedStyle(name='currency_style')
        currency_style.number_format = '"$"#,##0'
        currency_style.alignment = Alignment(horizontal='right')
        wb.add_named_style(currency_style)

        # Estilo de total
        total_style = NamedStyle(name='total_style')
        total_style.font = Font(bold=True)
        total_style.fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
        total_style.number_format = '"$"#,##0'
        wb.add_named_style(total_style)

    def _crear_hoja_resumen(self, wb: Workbook):
        """Crea la hoja de resumen ejecutivo."""
        ws = wb.active
        ws.title = "Resumen"

        # Título
        ws.merge_cells('A1:F1')
        ws['A1'] = "CUADRO DE COSTOS - MANTENIMIENTO LÍNEAS DE TRANSMISIÓN"
        ws['A1'].font = Font(bold=True, size=16)

        ws.merge_cells('A2:F2')
        ws['A2'] = f"Cliente: {self.linea.get_cliente_display() if self.linea else 'TODOS'}"

        ws.merge_cells('A3:F3')
        ws['A3'] = f"Período: {self.data['periodo']}"

        # Resumen de actividad
        row = 5
        ws[f'A{row}'] = "RESUMEN DE ACTIVIDAD"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        resumen_items = [
            ("Actividades Completadas", self.data['resumen']['actividades_completadas']),
            ("Torres Intervenidas", self.data['resumen']['torres_intervenidas']),
            ("Días Trabajados", self.data['resumen']['dias_trabajados']),
        ]

        for label, value in resumen_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1

        # Resumen de costos
        row += 1
        ws[f'A{row}'] = "RESUMEN DE COSTOS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        headers = ['Categoría', 'Subtotal']
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header).style = 'header_style'
        row += 1

        costos_items = [
            ("Personal", self.data['totales'].get('personal', 0)),
            ("Vehículos", self.data['totales'].get('vehiculos', 0)),
            ("Materiales", self.data['totales'].get('materiales', 0)),
            ("Herramientas", self.data['totales'].get('herramientas', 0)),
            ("Otros", self.data['totales'].get('otros', 0)),
        ]

        for label, value in costos_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = float(value)
            ws[f'B{row}'].style = 'currency_style'
            row += 1

        # Total
        ws[f'A{row}'] = "TOTAL"
        ws[f'B{row}'] = float(self.data['totales'].get('total', 0))
        ws[f'A{row}'].style = 'total_style'
        ws[f'B{row}'].style = 'total_style'

        # Ajustar anchos
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

    def _crear_hoja_personal(self, wb: Workbook):
        """Crea la hoja de costos de personal."""
        ws = wb.create_sheet("Personal")

        headers = ['Cargo', 'Nombre', 'Cuadrilla', 'Días', 'Valor Día', 'Subtotal']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).style = 'header_style'

        row = 2
        for item in self.data['personal']['detalle']:
            ws.cell(row=row, column=1, value=item['cargo'])
            ws.cell(row=row, column=2, value=item['nombre'])
            ws.cell(row=row, column=3, value=item['cuadrilla'])
            ws.cell(row=row, column=4, value=item['dias'])
            ws.cell(row=row, column=5, value=float(item['valor_dia'])).style = 'currency_style'
            ws.cell(row=row, column=6, value=float(item['subtotal'])).style = 'currency_style'
            row += 1

        # Total
        ws.cell(row=row, column=5, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=6, value=float(self.data['personal']['total'])).style = 'total_style'

        # Ajustar anchos
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _crear_hoja_vehiculos(self, wb: Workbook):
        """Crea la hoja de costos de vehículos."""
        ws = wb.create_sheet("Vehículos")

        headers = ['Placa', 'Tipo', 'Marca/Modelo', 'Cuadrilla', 'Días', 'Valor Día', 'Subtotal']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).style = 'header_style'

        row = 2
        for item in self.data['vehiculos']['detalle']:
            ws.cell(row=row, column=1, value=item['placa'])
            ws.cell(row=row, column=2, value=item['tipo'])
            ws.cell(row=row, column=3, value=item['marca_modelo'])
            ws.cell(row=row, column=4, value=item['cuadrilla'])
            ws.cell(row=row, column=5, value=item['dias'])
            ws.cell(row=row, column=6, value=float(item['valor_dia'])).style = 'currency_style'
            ws.cell(row=row, column=7, value=float(item['subtotal'])).style = 'currency_style'
            row += 1

        # Total
        ws.cell(row=row, column=6, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=7, value=float(self.data['vehiculos']['total'])).style = 'total_style'

    def _crear_hoja_materiales(self, wb: Workbook):
        """Crea la hoja de costos de materiales."""
        ws = wb.create_sheet("Materiales")

        headers = ['Código', 'Descripción', 'Unidad', 'Cantidad', 'Valor Unitario', 'Subtotal']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).style = 'header_style'

        row = 2
        for item in self.data['materiales']['detalle']:
            ws.cell(row=row, column=1, value=item['codigo'])
            ws.cell(row=row, column=2, value=item['descripcion'])
            ws.cell(row=row, column=3, value=item['unidad'])
            ws.cell(row=row, column=4, value=item['cantidad'])
            ws.cell(row=row, column=5, value=float(item['valor_unitario'])).style = 'currency_style'
            ws.cell(row=row, column=6, value=float(item['subtotal'])).style = 'currency_style'
            row += 1

        # Total
        ws.cell(row=row, column=5, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=6, value=float(self.data['materiales']['total'])).style = 'total_style'

    def _crear_hoja_detalle(self, wb: Workbook):
        """Crea la hoja de detalle por actividad."""
        ws = wb.create_sheet("Detalle Actividades")

        headers = ['Fecha', 'Línea', 'Torre', 'Tipo Actividad', 'Cuadrilla', 'Costo Total']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).style = 'header_style'

        # Ajustar anchos
        widths = [12, 15, 10, 30, 20, 15]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

    # Métodos de cálculo
    def _get_mes_nombre(self, mes: int) -> str:
        meses = [
            '', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
            'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
        ]
        return meses[mes]

    def _calcular_dias_trabajados(self, actividades) -> int:
        return actividades.values('fecha_programada').distinct().count()

    def _calcular_costos_personal(self, actividades) -> dict:
        """Calcula costos de personal basado en cuadrillas usadas."""
        detalle = []
        total = Decimal('0')

        cuadrillas_usadas = actividades.values('cuadrilla').distinct()
        dias = self._calcular_dias_trabajados(actividades)

        for cuadrilla_data in cuadrillas_usadas:
            if not cuadrilla_data['cuadrilla']:
                continue

            cuadrilla = Cuadrilla.objects.get(id=cuadrilla_data['cuadrilla'])
            miembros = CuadrillaMiembro.objects.filter(cuadrilla=cuadrilla, activo=True)

            for miembro in miembros:
                # Valor día según rol (simplificado)
                valores_dia = {
                    'supervisor': Decimal('150000'),
                    'liniero': Decimal('100000'),
                    'auxiliar': Decimal('70000'),
                }
                valor_dia = valores_dia.get(miembro.rol_cuadrilla, Decimal('80000'))
                subtotal = valor_dia * dias

                detalle.append({
                    'cargo': miembro.rol_cuadrilla.title(),
                    'nombre': miembro.usuario.get_full_name(),
                    'cuadrilla': cuadrilla.nombre,
                    'dias': dias,
                    'valor_dia': valor_dia,
                    'subtotal': subtotal,
                })
                total += subtotal

        return {'detalle': detalle, 'total': total}

    def _calcular_costos_vehiculos(self, actividades) -> dict:
        """Calcula costos de vehículos."""
        detalle = []
        total = Decimal('0')

        cuadrillas_usadas = actividades.values('cuadrilla').distinct()
        dias = self._calcular_dias_trabajados(actividades)

        for cuadrilla_data in cuadrillas_usadas:
            if not cuadrilla_data['cuadrilla']:
                continue

            cuadrilla = Cuadrilla.objects.select_related('vehiculo').get(
                id=cuadrilla_data['cuadrilla']
            )

            if cuadrilla.vehiculo:
                vehiculo = cuadrilla.vehiculo
                subtotal = vehiculo.costo_dia * dias

                detalle.append({
                    'placa': vehiculo.placa,
                    'tipo': vehiculo.get_tipo_display(),
                    'marca_modelo': f"{vehiculo.marca} {vehiculo.modelo}",
                    'cuadrilla': cuadrilla.nombre,
                    'dias': dias,
                    'valor_dia': vehiculo.costo_dia,
                    'subtotal': subtotal,
                })
                total += subtotal

        return {'detalle': detalle, 'total': total}

    def _calcular_costos_materiales(self, actividades) -> dict:
        """Calcula costos de materiales (simplificado)."""
        # En producción, obtener de modelo de consumo de materiales
        return {'detalle': [], 'total': Decimal('0')}

    def _calcular_costos_herramientas(self, actividades) -> dict:
        """Calcula costos de herramientas (simplificado)."""
        return {'detalle': [], 'total': Decimal('0')}

    def _calcular_otros_costos(self, actividades) -> dict:
        """Calcula otros costos (viáticos, etc.)."""
        return {'detalle': [], 'total': Decimal('0')}

    def _calcular_totales(self):
        """Calcula totales generales."""
        self.data['totales'] = {
            'personal': self.data['personal']['total'],
            'vehiculos': self.data['vehiculos']['total'],
            'materiales': self.data['materiales']['total'],
            'herramientas': self.data['herramientas']['total'],
            'otros': self.data['otros']['total'],
        }

        self.data['totales']['total'] = sum(self.data['totales'].values())
