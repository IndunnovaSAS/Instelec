"""
Views for crew management.
"""
from django.db import models
from django.shortcuts import redirect
from django.contrib import messages
from django.views.generic import ListView, DetailView, TemplateView
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse, JsonResponse
from apps.core.mixins import HTMXMixin, RoleRequiredMixin
from .models import Asistencia, Cuadrilla, CuadrillaMiembro, Vehiculo, TrackingUbicacion


class CuadrillaListView(LoginRequiredMixin, RoleRequiredMixin, HTMXMixin, ListView):
    """List all crews, organized by week."""
    model = Cuadrilla
    template_name = 'cuadrillas/lista.html'
    partial_template_name = 'cuadrillas/partials/lista_cuadrillas.html'
    context_object_name = 'cuadrillas'
    paginate_by = None
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente', 'supervisor']

    @staticmethod
    def _parse_semana(codigo):
        """Extract (week, year) from code format WW-YYYY-XXX."""
        try:
            parts = codigo.split('-')
            if len(parts) >= 2:
                semana = int(parts[0])
                ano = int(parts[1])
                if 1 <= semana <= 53 and 2000 <= ano <= 2100:
                    return semana, ano
        except (ValueError, IndexError):
            pass
        return None, None

    def get_queryset(self):
        qs = Cuadrilla.objects.filter(activa=True).select_related(
            'supervisor', 'vehiculo', 'linea_asignada'
        ).prefetch_related('miembros__usuario')

        # Filter by week if parameter provided
        semana_param = self.request.GET.get('semana', '').strip()
        if semana_param:
            # Format: WW-YYYY
            try:
                parts = semana_param.split('-')
                sem = parts[0].zfill(2)
                ano = parts[1]
                qs = qs.filter(codigo__startswith=f'{sem}-{ano}-')
            except (IndexError, ValueError):
                pass

        return qs

    def get_context_data(self, **kwargs):
        import json
        from collections import OrderedDict
        context = super().get_context_data(**kwargs)

        # Build list of available weeks from all active cuadrillas
        todas = Cuadrilla.objects.filter(activa=True).values_list('codigo', flat=True)
        semanas_set = set()
        for codigo in todas:
            sem, ano = self._parse_semana(codigo)
            if sem is not None:
                semanas_set.add((ano, sem))

        # Sort descending: most recent first
        semanas_disponibles = sorted(semanas_set, reverse=True)
        context['semanas_disponibles'] = [
            {'value': f'{s[1]}-{s[0]}', 'label': f'Semana {s[1]} - {s[0]}'}
            for s in semanas_disponibles
        ]

        # Current filter
        semana_param = self.request.GET.get('semana', '').strip()
        context['semana_actual'] = semana_param

        # Group cuadrillas by week for display
        cuadrillas_por_semana = OrderedDict()
        sin_semana = []
        for cuadrilla in context['cuadrillas']:
            sem, ano = self._parse_semana(cuadrilla.codigo)
            if sem is not None:
                key = f'Semana {sem} - {ano}'
                cuadrillas_por_semana.setdefault(key, []).append(cuadrilla)
            else:
                sin_semana.append(cuadrilla)

        if sin_semana:
            cuadrillas_por_semana['Otras'] = sin_semana

        context['cuadrillas_por_semana'] = cuadrillas_por_semana

        # Stats
        all_active = Cuadrilla.objects.filter(activa=True)
        context['total_cuadrillas'] = all_active.count()
        context['cuadrillas_activas'] = all_active.count()

        # Get latest location for each active crew for the mini-map
        ubicaciones = []
        for cuadrilla in context['cuadrillas']:
            ultima = TrackingUbicacion.objects.filter(
                cuadrilla=cuadrilla
            ).order_by('-created_at').first()

            if ultima:
                ubicaciones.append({
                    'cuadrilla_id': str(cuadrilla.id),
                    'cuadrilla_codigo': cuadrilla.codigo,
                    'lat': float(ultima.latitud),
                    'lng': float(ultima.longitud),
                })

        context['cuadrillas_ubicaciones_json'] = json.dumps(ubicaciones)
        return context


class CuadrillaDetailView(LoginRequiredMixin, RoleRequiredMixin, HTMXMixin, DetailView):
    """Detail view for a crew."""
    model = Cuadrilla
    template_name = 'cuadrillas/detalle.html'
    context_object_name = 'cuadrilla'
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente', 'supervisor']

    def get_context_data(self, **kwargs):
        from datetime import date, timedelta
        from decimal import Decimal
        from apps.usuarios.models import Usuario

        context = super().get_context_data(**kwargs)
        miembros = self.object.miembros.filter(activo=True).select_related('usuario')
        context['miembros'] = miembros

        # Total daily cost
        context['costo_total_dia'] = sum(
            (m.costo_dia for m in miembros), Decimal('0')
        )

        # Available users for add member form
        miembros_ids = miembros.values_list('usuario_id', flat=True)
        context['usuarios_disponibles'] = Usuario.objects.filter(
            is_active=True
        ).exclude(id__in=miembros_ids).order_by('first_name', 'last_name')

        # Choices for form selects
        context['roles_cuadrilla'] = CuadrillaMiembro.RolCuadrilla.choices
        context['cargos_jerarquicos'] = CuadrillaMiembro.CargoJerarquico.choices

        # Last known location
        ultima_ubicacion = TrackingUbicacion.objects.filter(
            cuadrilla=self.object
        ).order_by('-created_at').first()
        context['ultima_ubicacion'] = ultima_ubicacion

        # Weekly attendance calendar
        semana, ano = self._get_semana_from_codigo(self.object.codigo)
        if semana and ano:
            # ISO week: Monday=0 to Sunday=6
            lunes = date.fromisocalendar(ano, semana, 1)
        else:
            # Fallback: current week
            hoy = date.today()
            lunes = hoy - timedelta(days=hoy.weekday())

        dias_semana = [lunes + timedelta(days=i) for i in range(7)]
        context['dias_semana'] = dias_semana
        context['semana_lunes'] = lunes

        # Load existing attendance for this week
        asistencias = Asistencia.objects.filter(
            cuadrilla=self.object,
            fecha__in=dias_semana,
        ).select_related('usuario')

        # Build dict: {usuario_id: {fecha_iso: {tipo_novedad, viaticos, horas_extra, observacion, viatico_aplica}}}
        asistencia_por_usuario = {}
        for a in asistencias:
            uid = str(a.usuario_id)
            if uid not in asistencia_por_usuario:
                asistencia_por_usuario[uid] = {}
            asistencia_por_usuario[uid][a.fecha.isoformat()] = {
                'tipo_novedad': a.tipo_novedad,
                'viaticos': float(a.viaticos),
                'horas_extra': float(a.horas_extra),
                'observacion': a.observacion,
                'viatico_aplica': a.viatico_aplica,
            }

        # Build template-friendly structure
        filas_asistencia = []
        for miembro in miembros:
            uid = str(miembro.usuario_id)
            usuario_asistencia = asistencia_por_usuario.get(uid, {})
            dias = []
            total_viaticos = Decimal('0')
            total_horas_extra = Decimal('0')
            for dia in dias_semana:
                fecha_iso = dia.isoformat()
                info = usuario_asistencia.get(fecha_iso, {})
                viaticos_val = info.get('viaticos', 0)
                horas_extra_val = info.get('horas_extra', 0)
                total_viaticos += Decimal(str(viaticos_val))
                total_horas_extra += Decimal(str(horas_extra_val))
                dias.append({
                    'fecha': fecha_iso,
                    'fecha_display': dia.strftime('%d/%m'),
                    'tipo_novedad': info.get('tipo_novedad', ''),
                    'viaticos': viaticos_val,
                    'horas_extra': horas_extra_val,
                    'observacion': info.get('observacion', ''),
                    'viatico_aplica': info.get('viatico_aplica', False),
                })
            filas_asistencia.append({
                'miembro': miembro,
                'dias': dias,
                'total_viaticos': total_viaticos,
                'total_horas_extra': total_horas_extra,
            })

        context['filas_asistencia'] = filas_asistencia
        context['tipos_novedad'] = Asistencia.TipoNovedad.choices

        return context

    @staticmethod
    def _get_semana_from_codigo(codigo):
        """Extract (week, year) from code format WW-YYYY-XXX."""
        try:
            parts = codigo.split('-')
            if len(parts) >= 2:
                semana = int(parts[0])
                ano = int(parts[1])
                if 1 <= semana <= 53 and 2000 <= ano <= 2100:
                    return semana, ano
        except (ValueError, IndexError):
            pass
        return None, None


class CuadrillaEditView(LoginRequiredMixin, RoleRequiredMixin, HTMXMixin, DetailView):
    """View for editing a crew."""
    model = Cuadrilla
    template_name = 'cuadrillas/editar.html'
    context_object_name = 'cuadrilla'
    allowed_roles = ['admin', 'director', 'coordinador']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from apps.usuarios.models import Usuario
        from apps.lineas.models import Linea
        context['supervisores'] = Usuario.objects.filter(rol='supervisor', is_active=True)
        context['lineas'] = Linea.objects.filter(activa=True)
        context['vehiculos'] = Vehiculo.objects.filter(activo=True)
        return context

    def post(self, request, *args, **kwargs):
        """Handle form submission to update a crew."""
        from apps.usuarios.models import Usuario
        from apps.lineas.models import Linea

        cuadrilla = self.get_object()

        codigo = request.POST.get('codigo', '').strip()
        nombre = request.POST.get('nombre', '').strip()

        if not codigo or not nombre:
            messages.error(request, 'Código y nombre son obligatorios.')
            return self.get(request, *args, **kwargs)

        if Cuadrilla.objects.filter(codigo=codigo).exclude(pk=cuadrilla.pk).exists():
            messages.error(request, f'Ya existe otra cuadrilla con el código {codigo}.')
            return self.get(request, *args, **kwargs)

        try:
            cuadrilla.codigo = codigo
            cuadrilla.nombre = nombre

            supervisor_id = request.POST.get('supervisor') or None
            cuadrilla.supervisor = Usuario.objects.get(pk=supervisor_id) if supervisor_id else None

            vehiculo_id = request.POST.get('vehiculo') or None
            cuadrilla.vehiculo = Vehiculo.objects.get(pk=vehiculo_id) if vehiculo_id else None

            linea_id = request.POST.get('linea_asignada') or None
            cuadrilla.linea_asignada = Linea.objects.get(pk=linea_id) if linea_id else None

            fecha_str = request.POST.get('fecha', '').strip()
            if fecha_str:
                from datetime import date as date_cls
                cuadrilla.fecha = date_cls.fromisoformat(fecha_str)
            else:
                cuadrilla.fecha = None

            cuadrilla.activa = request.POST.get('activa') == 'on'
            cuadrilla.observaciones = request.POST.get('observaciones', '').strip()
            cuadrilla.save()
            messages.success(request, f'Cuadrilla {cuadrilla.codigo} actualizada exitosamente.')
            return redirect('cuadrillas:detalle', pk=cuadrilla.pk)
        except (Usuario.DoesNotExist, Linea.DoesNotExist, Vehiculo.DoesNotExist) as e:
            messages.error(request, f'Referencia inválida: {str(e)}')
            return self.get(request, *args, **kwargs)
        except Exception as e:
            messages.error(request, f'Error al actualizar la cuadrilla: {str(e)}')
            return self.get(request, *args, **kwargs)


class MapaCuadrillasView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    """Real-time map of all crews."""
    template_name = 'cuadrillas/mapa.html'
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente', 'supervisor']


class MapaCuadrillasPartialView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    """Partial view for HTMX polling of crew locations."""
    template_name = 'cuadrillas/partials/mapa_cuadrillas.html'
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente', 'supervisor']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Get latest location for each active crew
        cuadrillas = Cuadrilla.objects.filter(activa=True)
        ubicaciones = []

        for cuadrilla in cuadrillas:
            ultima = TrackingUbicacion.objects.filter(
                cuadrilla=cuadrilla
            ).order_by('-created_at').first()

            if ultima:
                ubicaciones.append({
                    'cuadrilla_id': str(cuadrilla.id),
                    'cuadrilla_codigo': cuadrilla.codigo,
                    'cuadrilla_nombre': cuadrilla.nombre,
                    'lat': float(ultima.latitud),
                    'lng': float(ultima.longitud),
                    'precision': float(ultima.precision_metros) if ultima.precision_metros else None,
                    'timestamp': ultima.created_at.isoformat(),
                })

        context['ubicaciones'] = ubicaciones
        return context

    def render_to_response(self, context, **response_kwargs):
        if self.request.headers.get('Accept') == 'application/json':
            return JsonResponse({'ubicaciones': context['ubicaciones']})
        return super().render_to_response(context, **response_kwargs)


class CuadrillaCreateView(LoginRequiredMixin, RoleRequiredMixin, HTMXMixin, TemplateView):
    """View for creating a new crew."""
    template_name = 'cuadrillas/crear.html'
    partial_template_name = 'cuadrillas/partials/form_cuadrilla.html'
    allowed_roles = ['admin', 'director', 'coordinador']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from apps.usuarios.models import Usuario
        from apps.lineas.models import Linea
        context['supervisores'] = Usuario.objects.filter(rol='supervisor', is_active=True)
        context['lineas'] = Linea.objects.filter(activa=True)
        return context

    def post(self, request, *args, **kwargs):
        """Handle form submission to create a new crew."""
        from django.shortcuts import redirect
        from django.contrib import messages
        from apps.usuarios.models import Usuario
        from apps.lineas.models import Linea

        codigo = request.POST.get('codigo', '').strip()
        nombre = request.POST.get('nombre', '').strip()
        supervisor_id = request.POST.get('supervisor') or None
        linea_id = request.POST.get('linea_asignada') or None

        # Validation
        if not codigo or not nombre:
            messages.error(request, 'Código y nombre son obligatorios.')
            return self.get(request, *args, **kwargs)

        if Cuadrilla.objects.filter(codigo=codigo).exists():
            messages.error(request, f'Ya existe una cuadrilla con el código {codigo}.')
            return self.get(request, *args, **kwargs)

        # Get related objects
        supervisor = None
        if supervisor_id:
            try:
                supervisor = Usuario.objects.get(pk=supervisor_id)
            except Usuario.DoesNotExist:
                pass

        linea_asignada = None
        if linea_id:
            try:
                linea_asignada = Linea.objects.get(pk=linea_id)
            except Linea.DoesNotExist:
                pass

        # Create the crew
        try:
            cuadrilla = Cuadrilla.objects.create(
                codigo=codigo,
                nombre=nombre,
                supervisor=supervisor,
                linea_asignada=linea_asignada,
            )
            messages.success(request, f'Cuadrilla {cuadrilla.codigo} creada exitosamente.')
            return redirect('cuadrillas:detalle', pk=cuadrilla.pk)
        except Exception as e:
            messages.error(request, f'Error al crear la cuadrilla: {str(e)}')
            return self.get(request, *args, **kwargs)


class CuadrillaMiembroAddView(LoginRequiredMixin, RoleRequiredMixin, DetailView):
    """Add a member to a cuadrilla."""
    model = Cuadrilla
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente', 'supervisor']

    def post(self, request, *args, **kwargs):
        from apps.usuarios.models import Usuario
        from datetime import date

        cuadrilla = self.get_object()
        usuario_id = request.POST.get('usuario')
        rol = request.POST.get('rol_cuadrilla', 'LINIERO_I')
        cargo = request.POST.get('cargo', 'MIEMBRO')
        costo_dia = request.POST.get('costo_dia', '0') or '0'

        if not usuario_id:
            messages.error(request, 'Debe seleccionar un usuario.')
            return redirect('cuadrillas:detalle', pk=cuadrilla.pk)

        try:
            usuario = Usuario.objects.get(pk=usuario_id)
        except Usuario.DoesNotExist:
            messages.error(request, 'Usuario no encontrado.')
            return redirect('cuadrillas:detalle', pk=cuadrilla.pk)

        if CuadrillaMiembro.objects.filter(
            cuadrilla=cuadrilla, usuario=usuario, activo=True
        ).exists():
            messages.error(request, f'{usuario.get_full_name()} ya es miembro activo.')
            return redirect('cuadrillas:detalle', pk=cuadrilla.pk)

        try:
            CuadrillaMiembro.objects.create(
                cuadrilla=cuadrilla,
                usuario=usuario,
                rol_cuadrilla=rol if rol in dict(CuadrillaMiembro.RolCuadrilla.choices) else 'LINIERO_I',
                cargo=cargo if cargo in dict(CuadrillaMiembro.CargoJerarquico.choices) else 'MIEMBRO',
                costo_dia=float(costo_dia),
                fecha_inicio=date.today(),
            )
            messages.success(request, f'{usuario.get_full_name()} agregado a la cuadrilla.')
        except Exception as e:
            messages.error(request, f'Error al agregar miembro: {str(e)}')

        return redirect('cuadrillas:detalle', pk=cuadrilla.pk)


class AsistenciaUpdateView(LoginRequiredMixin, RoleRequiredMixin, DetailView):
    """Update attendance for a crew member on a specific day (HTMX)."""
    model = Cuadrilla
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente', 'supervisor']

    def post(self, request, *args, **kwargs):
        from datetime import date as date_cls
        from decimal import Decimal, InvalidOperation
        from django.http import HttpResponse

        cuadrilla = self.get_object()
        usuario_id = request.POST.get('usuario_id', '').strip()
        fecha_str = request.POST.get('fecha', '').strip()
        tipo_novedad = request.POST.get('tipo_novedad', '').strip()
        viaticos_str = request.POST.get('viaticos', '').strip()
        horas_extra_str = request.POST.get('horas_extra', '').strip()
        observacion = request.POST.get('observacion', '').strip()
        viatico_aplica = request.POST.get('viatico_aplica') == 'on'

        if not usuario_id or not fecha_str:
            return HttpResponse('Datos incompletos', status=400)

        try:
            fecha = date_cls.fromisoformat(fecha_str)
        except ValueError:
            return HttpResponse('Fecha invalida', status=400)

        # Verify user is member of the cuadrilla
        if not CuadrillaMiembro.objects.filter(
            cuadrilla=cuadrilla, usuario_id=usuario_id, activo=True
        ).exists():
            return HttpResponse('Usuario no es miembro', status=400)

        # Parse viaticos
        try:
            viaticos = Decimal(viaticos_str) if viaticos_str else Decimal('0')
        except InvalidOperation:
            viaticos = Decimal('0')

        # Parse horas_extra
        try:
            horas_extra = Decimal(horas_extra_str) if horas_extra_str else Decimal('0')
        except InvalidOperation:
            horas_extra = Decimal('0')

        # If viatico_aplica, calculate viaticos from CostoRecurso
        if viatico_aplica:
            from apps.financiero.models import CostoRecurso
            costo_viatico = CostoRecurso.objects.filter(
                tipo='VIATICO', activo=True
            ).order_by('-vigencia_desde').first()
            if costo_viatico:
                viaticos = costo_viatico.costo_unitario
            else:
                viaticos = Decimal('136941')
        elif not viaticos_str:
            viaticos = Decimal('0')

        tipos_validos = dict(Asistencia.TipoNovedad.choices)

        if not tipo_novedad and not viatico_aplica:
            # Empty selection and no viatico: remove attendance record
            Asistencia.objects.filter(
                usuario_id=usuario_id, cuadrilla=cuadrilla, fecha=fecha
            ).delete()
            viaticos = Decimal('0')
        elif not tipo_novedad and viatico_aplica:
            # No novedad but viatico checked: save with PRESENTE default
            Asistencia.objects.update_or_create(
                usuario_id=usuario_id,
                cuadrilla=cuadrilla,
                fecha=fecha,
                defaults={
                    'tipo_novedad': Asistencia.TipoNovedad.PRESENTE,
                    'viaticos': viaticos,
                    'horas_extra': horas_extra,
                    'observacion': observacion,
                    'viatico_aplica': viatico_aplica,
                    'registrado_por': request.user,
                }
            )
            tipo_novedad = 'PRESENTE'
        elif tipo_novedad in tipos_validos:
            Asistencia.objects.update_or_create(
                usuario_id=usuario_id,
                cuadrilla=cuadrilla,
                fecha=fecha,
                defaults={
                    'tipo_novedad': tipo_novedad,
                    'viaticos': viaticos,
                    'horas_extra': horas_extra,
                    'observacion': observacion,
                    'viatico_aplica': viatico_aplica,
                    'registrado_por': request.user,
                }
            )
        else:
            return HttpResponse('Tipo de novedad invalido', status=400)

        # Return the updated cell content
        color_map = {
            'PRESENTE': 'text-green-600 bg-green-50 border-green-300',
            'AUSENTE': 'text-red-600 bg-red-50 border-red-300',
            'VACACIONES': 'text-blue-600 bg-blue-50 border-blue-300',
            'INCAPACIDAD': 'text-orange-600 bg-orange-50 border-orange-300',
            'PERMISO': 'text-purple-600 bg-purple-50 border-purple-300',
            'LICENCIA': 'text-yellow-700 bg-yellow-50 border-yellow-300',
            'CAPACITACION': 'text-teal-600 bg-teal-50 border-teal-300',
        }
        css = color_map.get(tipo_novedad, 'text-gray-400 bg-white border-gray-200')

        options_html = '<option value="">---</option>'
        for val, lbl in Asistencia.TipoNovedad.choices:
            sel = ' selected' if val == tipo_novedad else ''
            options_html += f'<option value="{val}"{sel}>{lbl}</option>'

        horas_extra_display = float(horas_extra) if horas_extra else ''
        viatico_checked = ' checked' if viatico_aplica else ''
        obs_escaped = observacion.replace('"', '&quot;')

        # Calculate weekly total for OOB swap
        from datetime import timedelta
        semana, ano = CuadrillaDetailView._get_semana_from_codigo(cuadrilla.codigo)
        if semana and ano:
            lunes = date_cls.fromisocalendar(ano, semana, 1)
        else:
            hoy = date_cls.today()
            lunes = hoy - timedelta(days=hoy.weekday())
        dias_semana = [lunes + timedelta(days=i) for i in range(7)]

        total_viaticos_semana = Asistencia.objects.filter(
            usuario_id=usuario_id,
            cuadrilla=cuadrilla,
            fecha__in=dias_semana,
        ).aggregate(total=models.Sum('viaticos'))['total'] or Decimal('0')
        total_viaticos_fmt = int(total_viaticos_semana)

        # Build observation field (visible when not PRESENTE)
        if tipo_novedad and tipo_novedad != 'PRESENTE':
            obs_field = (
                f'<input type="text" name="observacion" value="{obs_escaped}" '
                f'placeholder="Motivo de ausencia..." '
                f'hx-post="{request.path}" '
                f'hx-target="closest .asistencia-cell" '
                f'hx-swap="innerHTML" '
                f'hx-include="closest .asistencia-cell" '
                f'hx-trigger="change" '
                f'class="mt-1 text-xs rounded border border-yellow-300 bg-yellow-50 px-1 py-0.5 w-full '
                f'text-gray-700 dark:bg-gray-700 dark:border-gray-600 dark:text-gray-200">'
            )
        else:
            obs_field = f'<input type="hidden" name="observacion" value="{obs_escaped}">'

        html = (
            f'<select name="tipo_novedad" '
            f'hx-post="{request.path}" '
            f'hx-target="closest .asistencia-cell" '
            f'hx-swap="innerHTML" '
            f'hx-include="closest .asistencia-cell" '
            f'class="text-xs rounded border px-1 py-1 w-full cursor-pointer {css} dark:bg-gray-700 dark:border-gray-600 dark:text-gray-200">'
            f'{options_html}</select>'
            f'{obs_field}'
            f'<div class="mt-1 flex items-center gap-1">'
            f'<input type="checkbox" name="viatico_aplica" {viatico_checked} '
            f'hx-post="{request.path}" '
            f'hx-target="closest .asistencia-cell" '
            f'hx-swap="innerHTML" '
            f'hx-include="closest .asistencia-cell" '
            f'class="rounded border-gray-300 text-green-600 cursor-pointer">'
            f'<span class="text-xs text-gray-500">V</span>'
            f'</div>'
            f'<input type="number" name="horas_extra" value="{horas_extra_display}" step="0.5" min="0" '
            f'hx-post="{request.path}" '
            f'hx-target="closest .asistencia-cell" '
            f'hx-swap="innerHTML" '
            f'hx-include="closest .asistencia-cell" '
            f'hx-trigger="change" '
            f'class="mt-1 text-xs rounded border border-orange-200 bg-orange-50 px-1 py-0.5 w-full text-center '
            f'text-orange-700 placeholder-orange-300 dark:bg-gray-700 dark:border-gray-600 dark:text-orange-300 '
            f'dark:placeholder-orange-600" '
            f'placeholder="Horas Extra">'
            f'<input type="hidden" name="usuario_id" value="{usuario_id}">'
            f'<input type="hidden" name="fecha" value="{fecha_str}">'
            f'<input type="hidden" name="viaticos" value="{float(viaticos)}">'
            # OOB swap: update the total viaticos span for this user
            f'<span id="total-viaticos-{usuario_id}" hx-swap-oob="true" '
            f'class="text-sm font-bold text-green-600 dark:text-green-400">'
            f'${total_viaticos_fmt}'
            f'</span>'
        )
        return HttpResponse(html)


class CuadrillaMiembroRemoveView(LoginRequiredMixin, RoleRequiredMixin, DetailView):
    """Remove a member from a cuadrilla (soft delete)."""
    model = Cuadrilla
    allowed_roles = ['admin', 'director', 'coordinador']

    def post(self, request, *args, **kwargs):
        from datetime import date

        cuadrilla = self.get_object()
        miembro_pk = self.kwargs['miembro_pk']

        try:
            miembro = CuadrillaMiembro.objects.get(pk=miembro_pk, cuadrilla=cuadrilla)
            nombre = miembro.usuario.get_full_name()
            miembro.activo = False
            miembro.fecha_fin = date.today()
            miembro.save(update_fields=['activo', 'fecha_fin', 'updated_at'])
            messages.success(request, f'{nombre} removido de la cuadrilla.')
        except CuadrillaMiembro.DoesNotExist:
            messages.error(request, 'Miembro no encontrado.')

        return redirect('cuadrillas:detalle', pk=cuadrilla.pk)


class ExportarAsistenciaView(LoginRequiredMixin, RoleRequiredMixin, View):
    """Export weekly attendance to Excel."""
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente', 'supervisor']

    def get(self, request, pk, *args, **kwargs):
        from datetime import date, timedelta
        from decimal import Decimal
        from io import BytesIO

        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        try:
            cuadrilla = Cuadrilla.objects.get(pk=pk)
        except Cuadrilla.DoesNotExist:
            return HttpResponse('Cuadrilla no encontrada', status=404)

        # Determine week from cuadrilla code
        semana, ano = self._get_semana_from_codigo(cuadrilla.codigo)
        if semana and ano:
            lunes = date.fromisocalendar(ano, semana, 1)
        else:
            hoy = date.today()
            lunes = hoy - timedelta(days=hoy.weekday())

        dias_semana = [lunes + timedelta(days=i) for i in range(7)]
        dias_nombres = ['Lun', 'Mar', 'Mie', 'Jue', 'Vie', 'Sab', 'Dom']

        miembros = cuadrilla.miembros.filter(activo=True).select_related('usuario')
        asistencias = Asistencia.objects.filter(
            cuadrilla=cuadrilla,
            fecha__in=dias_semana,
        ).select_related('usuario')

        # Build dict
        asist_dict = {}
        for a in asistencias:
            uid = str(a.usuario_id)
            if uid not in asist_dict:
                asist_dict[uid] = {}
            asist_dict[uid][a.fecha.isoformat()] = a

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Asistencia'

        # Styles
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center = Alignment(horizontal='center', vertical='center')

        # Title row
        ws.merge_cells('A1:L1')
        ws['A1'] = f'Asistencia Semanal - {cuadrilla.codigo} - {cuadrilla.nombre}'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A2:L2')
        ws['A2'] = f'Semana: {dias_semana[0].strftime("%d/%m/%Y")} - {dias_semana[6].strftime("%d/%m/%Y")}'
        ws['A2'].font = Font(size=11)

        # Headers (row 4)
        headers = ['Nombre', 'Documento', 'Cargo', 'Rol']
        for dia, nombre in zip(dias_semana, dias_nombres):
            headers.append(f'{nombre} {dia.strftime("%d/%m")}')
        headers.extend(['Total Viaticos', 'H. Extra', 'Observaciones'])

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

        # Data rows
        novedad_labels = dict(Asistencia.TipoNovedad.choices)
        row = 5
        for miembro in miembros:
            uid = str(miembro.usuario_id)
            user_asist = asist_dict.get(uid, {})

            ws.cell(row=row, column=1, value=miembro.usuario.get_full_name()).border = thin_border
            ws.cell(row=row, column=2, value=getattr(miembro.usuario, 'documento', '')).border = thin_border
            ws.cell(row=row, column=3, value=miembro.get_rol_cuadrilla_display()).border = thin_border
            ws.cell(row=row, column=4, value=miembro.get_cargo_display()).border = thin_border

            total_viaticos = Decimal('0')
            total_horas_extra = Decimal('0')
            observaciones_semana = []

            for i, dia in enumerate(dias_semana):
                asist = user_asist.get(dia.isoformat())
                col = 5 + i
                if asist:
                    cell = ws.cell(row=row, column=col, value=novedad_labels.get(asist.tipo_novedad, ''))
                    cell.alignment = center
                    cell.border = thin_border
                    total_viaticos += asist.viaticos
                    total_horas_extra += asist.horas_extra
                    if asist.observacion:
                        observaciones_semana.append(f'{dias_nombres[i]}: {asist.observacion}')

                    # Color coding
                    color_map = {
                        'PRESENTE': '92D050',
                        'AUSENTE': 'FF6B6B',
                        'VACACIONES': '6BB5FF',
                        'INCAPACIDAD': 'FFB366',
                        'PERMISO': 'C39BD3',
                        'LICENCIA': 'F7DC6F',
                        'CAPACITACION': '76D7C4',
                    }
                    fill_color = color_map.get(asist.tipo_novedad)
                    if fill_color:
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                else:
                    cell = ws.cell(row=row, column=col, value='---')
                    cell.alignment = center
                    cell.border = thin_border

            # Totals
            cell_v = ws.cell(row=row, column=12, value=float(total_viaticos))
            cell_v.number_format = '$#,##0'
            cell_v.alignment = center
            cell_v.border = thin_border

            cell_h = ws.cell(row=row, column=13, value=float(total_horas_extra))
            cell_h.number_format = '0.0'
            cell_h.alignment = center
            cell_h.border = thin_border

            ws.cell(row=row, column=14, value='; '.join(observaciones_semana)).border = thin_border

            row += 1

        # Auto-fit column widths
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 35)

        # Write to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f'asistencia_{cuadrilla.codigo}_{lunes.strftime("%Y%m%d")}.xlsx'
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @staticmethod
    def _get_semana_from_codigo(codigo):
        try:
            parts = codigo.split('-')
            if len(parts) >= 2:
                semana = int(parts[0])
                ano = int(parts[1])
                if 1 <= semana <= 53 and 2000 <= ano <= 2100:
                    return semana, ano
        except (ValueError, IndexError):
            pass
        return None, None


class CostoRolAPIView(LoginRequiredMixin, RoleRequiredMixin, View):
    """API endpoint to get cost by role."""
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente', 'supervisor']

    def get(self, request, *args, **kwargs):
        rol = request.GET.get('rol', '').strip()
        if not rol:
            return JsonResponse({'costo_dia': 0})

        # Costos fijos por rol
        costos = {
            'SUPERVISOR': 0,
            'LINIERO_I': 3176095,
            'LINIERO_II': 2804856,
            'AYUDANTE': 1750905,
            'CONDUCTOR': 0,
            'ADMINISTRADOR_OBRA': 2522400,
            'PROFESIONAL_SST': 4204000,
            'ING_RESIDENTE': 7357000,
            'SERVICIO_GENERAL': 1750905,
            'ALMACENISTA': 1800000,
            'SUPERVISOR_FOREST': 2969427,
            'ASISTENTE_FOREST': 4204000,
        }
        return JsonResponse({'costo_dia': costos.get(rol, 0)})
