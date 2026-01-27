"""
Importers for activity programming from Excel files.
"""
import logging
from typing import Any
from decimal import Decimal
from datetime import date

from openpyxl import load_workbook
from django.db import transaction

logger = logging.getLogger(__name__)


class ProgramaTranselcaImporter:
    """
    Importa Excel de Transelca con columnas:
    - Aviso SAP
    - Línea
    - Tipo Actividad
    - Mes programado
    - Ejecutor (OUTSOURCING)
    - Tramo (opcional)
    - Torre inicio / Torre fin (opcional)
    - Valor facturación (opcional)

    El formato puede variar, por lo que se intenta detectar las columnas
    por sus nombres en la primera fila.
    """

    # Mapeo de nombres de columnas posibles a campos internos
    COLUMN_MAPPINGS = {
        'aviso_sap': ['aviso sap', 'aviso', 'nro aviso', 'numero aviso', 'sap', 'no. aviso'],
        'linea': ['línea', 'linea', 'line', 'codigo linea', 'código línea'],
        'tipo_actividad': ['tipo actividad', 'actividad', 'tipo', 'tipo de actividad', 'descripcion actividad'],
        'mes': ['mes', 'mes programado', 'fecha programada', 'mes ejecucion'],
        'ejecutor': ['ejecutor', 'contratista', 'outsourcing', 'empresa'],
        'tramo': ['tramo', 'sector', 'seccion'],
        'torre_inicio': ['torre inicio', 'torre ini', 'desde torre', 'torre desde'],
        'torre_fin': ['torre fin', 'torre final', 'hasta torre', 'torre hasta'],
        'valor_facturacion': ['valor', 'valor facturacion', 'facturacion', 'precio', 'monto'],
        'observaciones': ['observaciones', 'notas', 'comentarios', 'obs'],
    }

    def __init__(self):
        self.errores = []
        self.advertencias = []
        self.actividades_creadas = []
        self.actividades_actualizadas = []
        self.filas_omitidas = []
        self.column_indices = {}

    def importar(self, archivo_excel, programacion_mensual, opciones=None):
        """
        Importa actividades desde un archivo Excel de Transelca.

        Args:
            archivo_excel: File object or path to Excel file
            programacion_mensual: ProgramacionMensual instance to associate activities
            opciones: Dict with import options (actualizar_existentes, etc.)

        Returns:
            Dict with import summary
        """
        from .models import Actividad, TipoActividad
        from apps.lineas.models import Linea, Torre, Tramo

        opciones = opciones or {}
        actualizar_existentes = opciones.get('actualizar_existentes', False)

        try:
            workbook = load_workbook(archivo_excel, read_only=True, data_only=True)
            sheet = workbook.active
        except Exception as e:
            logger.error(f"Error loading Excel file: {e}")
            return {
                'exito': False,
                'error': f'Error al cargar archivo Excel: {str(e)}',
                'actividades_creadas': 0,
                'actividades_actualizadas': 0,
            }

        # Detectar columnas en la primera fila
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return {
                'exito': False,
                'error': 'El archivo está vacío',
                'actividades_creadas': 0,
                'actividades_actualizadas': 0,
            }

        header_row = rows[0]
        self._detectar_columnas(header_row)

        if 'linea' not in self.column_indices:
            return {
                'exito': False,
                'error': 'No se encontró la columna de Línea en el archivo',
                'actividades_creadas': 0,
                'actividades_actualizadas': 0,
            }

        if 'tipo_actividad' not in self.column_indices:
            return {
                'exito': False,
                'error': 'No se encontró la columna de Tipo de Actividad en el archivo',
                'actividades_creadas': 0,
                'actividades_actualizadas': 0,
            }

        # Procesar filas de datos
        linea_asociada = programacion_mensual.linea

        with transaction.atomic():
            for row_num, row in enumerate(rows[1:], start=2):
                try:
                    resultado = self._procesar_fila(
                        row, row_num, programacion_mensual, linea_asociada, actualizar_existentes
                    )
                    if resultado == 'creada':
                        self.actividades_creadas.append(row_num)
                    elif resultado == 'actualizada':
                        self.actividades_actualizadas.append(row_num)
                    elif resultado == 'omitida':
                        self.filas_omitidas.append(row_num)
                except Exception as e:
                    logger.warning(f"Error processing row {row_num}: {e}")
                    self.errores.append({
                        'fila': row_num,
                        'error': str(e)
                    })

        # Actualizar programación mensual
        programacion_mensual.total_actividades = Actividad.objects.filter(
            programacion=programacion_mensual
        ).count()
        programacion_mensual.save(update_fields=['total_actividades', 'updated_at'])

        return {
            'exito': True,
            'actividades_creadas': len(self.actividades_creadas),
            'actividades_actualizadas': len(self.actividades_actualizadas),
            'filas_omitidas': len(self.filas_omitidas),
            'errores': self.errores,
            'advertencias': self.advertencias,
            'columnas_detectadas': list(self.column_indices.keys()),
        }

    def _detectar_columnas(self, header_row):
        """Detecta las columnas en la fila de encabezado."""
        for col_idx, cell_value in enumerate(header_row):
            if cell_value is None:
                continue
            cell_lower = str(cell_value).lower().strip()

            for field_name, posibles in self.COLUMN_MAPPINGS.items():
                if cell_lower in posibles:
                    self.column_indices[field_name] = col_idx
                    break

        logger.info(f"Detected columns: {self.column_indices}")

    def _get_cell_value(self, row, field_name):
        """Obtiene el valor de una celda por nombre de campo."""
        if field_name not in self.column_indices:
            return None
        idx = self.column_indices[field_name]
        if idx < len(row):
            return row[idx]
        return None

    def _procesar_fila(self, row, row_num, programacion_mensual, linea_asociada, actualizar_existentes):
        """Procesa una fila del Excel y crea/actualiza la actividad."""
        from .models import Actividad, TipoActividad
        from apps.lineas.models import Linea, Torre, Tramo

        # Obtener valores de la fila
        aviso_sap = self._get_cell_value(row, 'aviso_sap')
        linea_codigo = self._get_cell_value(row, 'linea')
        tipo_actividad_nombre = self._get_cell_value(row, 'tipo_actividad')
        tramo_codigo = self._get_cell_value(row, 'tramo')
        torre_inicio_num = self._get_cell_value(row, 'torre_inicio')
        torre_fin_num = self._get_cell_value(row, 'torre_fin')
        valor_facturacion = self._get_cell_value(row, 'valor_facturacion')
        observaciones = self._get_cell_value(row, 'observaciones')

        # Validaciones básicas
        if not linea_codigo and not linea_asociada:
            self.advertencias.append({
                'fila': row_num,
                'mensaje': 'No se especificó línea y no hay línea asociada a la programación'
            })
            return 'omitida'

        if not tipo_actividad_nombre:
            self.advertencias.append({
                'fila': row_num,
                'mensaje': 'No se especificó tipo de actividad'
            })
            return 'omitida'

        # Buscar línea
        linea = linea_asociada
        if linea_codigo:
            try:
                linea = Linea.objects.get(codigo__iexact=str(linea_codigo).strip())
            except Linea.DoesNotExist:
                self.advertencias.append({
                    'fila': row_num,
                    'mensaje': f'Línea no encontrada: {linea_codigo}'
                })
                if not linea_asociada:
                    return 'omitida'

        # Buscar tipo de actividad
        try:
            tipo_actividad = TipoActividad.objects.get(
                nombre__iexact=str(tipo_actividad_nombre).strip()
            )
        except TipoActividad.DoesNotExist:
            # Intentar búsqueda parcial
            tipos = TipoActividad.objects.filter(
                nombre__icontains=str(tipo_actividad_nombre).strip()
            )
            if tipos.exists():
                tipo_actividad = tipos.first()
                self.advertencias.append({
                    'fila': row_num,
                    'mensaje': f'Tipo de actividad "{tipo_actividad_nombre}" mapeado a "{tipo_actividad.nombre}"'
                })
            else:
                self.advertencias.append({
                    'fila': row_num,
                    'mensaje': f'Tipo de actividad no encontrado: {tipo_actividad_nombre}'
                })
                return 'omitida'

        # Buscar tramo (opcional)
        tramo = None
        if tramo_codigo:
            try:
                tramo = Tramo.objects.get(codigo__iexact=str(tramo_codigo).strip())
            except Tramo.DoesNotExist:
                self.advertencias.append({
                    'fila': row_num,
                    'mensaje': f'Tramo no encontrado: {tramo_codigo}'
                })

        # Buscar torre (opcional, usa la torre de inicio del tramo si hay tramo)
        torre = None
        if tramo:
            torre = tramo.torre_inicio
        elif torre_inicio_num:
            try:
                torre = Torre.objects.get(
                    linea=linea,
                    numero=str(torre_inicio_num).strip()
                )
            except Torre.DoesNotExist:
                self.advertencias.append({
                    'fila': row_num,
                    'mensaje': f'Torre no encontrada: {torre_inicio_num}'
                })

        # Preparar valor de facturación
        valor_fact = Decimal('0')
        if valor_facturacion:
            try:
                valor_fact = Decimal(str(valor_facturacion).replace(',', '.').replace('$', '').strip())
            except (ValueError, TypeError):
                self.advertencias.append({
                    'fila': row_num,
                    'mensaje': f'Valor de facturación inválido: {valor_facturacion}'
                })

        # Verificar si ya existe (por aviso SAP)
        actividad_existente = None
        if aviso_sap:
            aviso_sap_str = str(aviso_sap).strip()
            try:
                actividad_existente = Actividad.objects.get(aviso_sap=aviso_sap_str)
            except Actividad.DoesNotExist:
                pass

        if actividad_existente:
            if actualizar_existentes:
                # Actualizar actividad existente
                actividad_existente.linea = linea
                actividad_existente.tipo_actividad = tipo_actividad
                actividad_existente.torre = torre
                actividad_existente.tramo = tramo
                actividad_existente.programacion = programacion_mensual
                if valor_fact > 0:
                    actividad_existente.valor_facturacion = valor_fact
                if observaciones:
                    actividad_existente.observaciones_programacion = str(observaciones)
                actividad_existente.save()
                return 'actualizada'
            else:
                self.advertencias.append({
                    'fila': row_num,
                    'mensaje': f'Actividad con Aviso SAP {aviso_sap} ya existe, omitiendo'
                })
                return 'omitida'

        # Crear nueva actividad
        actividad = Actividad.objects.create(
            linea=linea,
            torre=torre,
            tipo_actividad=tipo_actividad,
            programacion=programacion_mensual,
            tramo=tramo,
            aviso_sap=str(aviso_sap).strip() if aviso_sap else '',
            fecha_programada=date(programacion_mensual.anio, programacion_mensual.mes, 1),
            estado=Actividad.Estado.PENDIENTE,
            prioridad=Actividad.Prioridad.NORMAL,
            valor_facturacion=valor_fact,
            observaciones_programacion=str(observaciones) if observaciones else '',
        )

        return 'creada'


class ImportadorExcelGenerico:
    """
    Importador genérico para diferentes formatos de Excel.
    Útil para importar datos de otras fuentes.
    """

    def __init__(self, mapping_columnas=None):
        self.mapping_columnas = mapping_columnas or {}
        self.errores = []
        self.registros_procesados = 0

    def leer_excel(self, archivo_excel, hoja=None):
        """
        Lee un archivo Excel y retorna los datos como lista de diccionarios.

        Args:
            archivo_excel: File object or path
            hoja: Nombre de la hoja (None = hoja activa)

        Returns:
            List of dicts with row data
        """
        try:
            workbook = load_workbook(archivo_excel, read_only=True, data_only=True)
            if hoja:
                sheet = workbook[hoja]
            else:
                sheet = workbook.active
        except Exception as e:
            logger.error(f"Error loading Excel: {e}")
            return []

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(h).strip() if h else f'col_{i}' for i, h in enumerate(rows[0])]
        data = []

        for row in rows[1:]:
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    col_name = headers[i]
                    # Aplicar mapping si existe
                    if col_name.lower() in self.mapping_columnas:
                        col_name = self.mapping_columnas[col_name.lower()]
                    row_dict[col_name] = value
            data.append(row_dict)

        return data
