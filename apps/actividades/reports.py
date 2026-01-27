"""
Advanced reports for activity tracking.
"""
import logging
from io import BytesIO
from datetime import date, timedelta
from decimal import Decimal
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ReporteAvanceServidumbre:
    """
    Genera Excel con:
    - Vanos ejecutados (verde oscuro = completo, verde claro = parcial, azul = pendiente)
    - Lista de pendientes con descripción
    - Gráfico de torta de avance
    """

    # Colores
    COLOR_COMPLETO = PatternFill(start_color='006400', end_color='006400', fill_type='solid')
    COLOR_PARCIAL = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    COLOR_PENDIENTE = PatternFill(start_color='4169E1', end_color='4169E1', fill_type='solid')
    COLOR_CON_CONDICION = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

    HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
    CELL_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def __init__(self):
        self.workbook = None

    def generar(self, linea_id, tramo_id=None, fecha_corte=None):
        """
        Genera el reporte de avance de servidumbre.

        Args:
            linea_id: UUID de la línea
            tramo_id: UUID del tramo (opcional, para filtrar)
            fecha_corte: Fecha de corte para el reporte

        Returns:
            BytesIO con el archivo Excel
        """
        from apps.lineas.models import Linea, Torre, Tramo
        from apps.actividades.models import Actividad
        from apps.campo.models import RegistroCampo

        if fecha_corte is None:
            fecha_corte = date.today()

        try:
            linea = Linea.objects.get(id=linea_id)
        except Linea.DoesNotExist:
            raise ValueError(f'Línea no encontrada: {linea_id}')

        self.workbook = Workbook()

        # Obtener torres (filtrar por tramo si se especifica)
        if tramo_id:
            try:
                tramo = Tramo.objects.get(id=tramo_id)
                torres = tramo.torres_incluidas
            except Tramo.DoesNotExist:
                torres = Torre.objects.filter(linea=linea)
        else:
            torres = Torre.objects.filter(linea=linea)

        torres = torres.order_by('numero')

        # Generar hojas
        self._generar_hoja_vanos(linea, torres, fecha_corte)
        self._generar_hoja_pendientes(linea, torres)
        self._generar_hoja_resumen_con_grafico(linea, torres, fecha_corte)

        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)

        return output

    def _generar_hoja_vanos(self, linea, torres, fecha_corte):
        """Genera matriz visual de vanos con colores."""
        from apps.actividades.models import Actividad

        sheet = self.workbook.active
        sheet.title = 'Estado Vanos'

        # Título
        sheet.merge_cells('A1:L1')
        titulo = sheet['A1']
        titulo.value = f'ESTADO DE VANOS - {linea.codigo}'
        titulo.font = Font(bold=True, size=14)
        titulo.alignment = Alignment(horizontal='center')

        sheet['A2'] = f'Fecha de corte: {fecha_corte.strftime("%d/%m/%Y")}'
        sheet['A3'] = 'Leyenda: '
        sheet['B3'] = 'Completo (100%)'
        sheet['B3'].fill = self.COLOR_COMPLETO
        sheet['B3'].font = Font(color='FFFFFF')
        sheet['D3'] = 'Parcial (1-99%)'
        sheet['D3'].fill = self.COLOR_PARCIAL
        sheet['F3'] = 'Pendiente (0%)'
        sheet['F3'].fill = self.COLOR_PENDIENTE
        sheet['F3'].font = Font(color='FFFFFF')
        sheet['H3'] = 'Con Condición'
        sheet['H3'].fill = self.COLOR_CON_CONDICION

        # Crear matriz de vanos
        # Cada fila representa un rango de 10 vanos
        row_num = 5
        torres_list = list(torres)
        vanos_por_fila = 10

        # Encabezado de columnas
        sheet.cell(row=row_num, column=1, value='Vano').font = Font(bold=True)
        for i in range(vanos_por_fila):
            cell = sheet.cell(row=row_num, column=i+2, value=f'+{i}')
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        row_num += 1

        # Procesar vanos
        for i in range(0, len(torres_list) - 1, vanos_por_fila):
            # Etiqueta de fila (rango de vanos)
            vano_inicio = i + 1
            vano_fin = min(i + vanos_por_fila, len(torres_list) - 1)
            sheet.cell(row=row_num, column=1, value=f'V{vano_inicio}-V{vano_fin}')

            for j in range(vanos_por_fila):
                vano_idx = i + j
                if vano_idx >= len(torres_list) - 1:
                    break

                torre_inicio = torres_list[vano_idx]
                torre_fin = torres_list[vano_idx + 1]

                # Obtener actividades para este vano
                actividades = Actividad.objects.filter(
                    linea=linea,
                    torre__in=[torre_inicio, torre_fin]
                )

                # Calcular estado del vano
                if actividades.exists():
                    avance_total = sum(a.porcentaje_avance for a in actividades)
                    avance_promedio = avance_total / actividades.count()

                    # Verificar si hay condiciones/pendientes
                    tiene_condicion = any(
                        r.tiene_pendiente
                        for a in actividades
                        for r in a.registros_campo.all()
                    )
                else:
                    avance_promedio = 0
                    tiene_condicion = False

                # Determinar color
                cell = sheet.cell(row=row_num, column=j+2)
                cell.value = f'{avance_promedio:.0f}%'
                cell.alignment = Alignment(horizontal='center')
                cell.border = self.CELL_BORDER

                if tiene_condicion:
                    cell.fill = self.COLOR_CON_CONDICION
                elif avance_promedio >= 100:
                    cell.fill = self.COLOR_COMPLETO
                    cell.font = Font(color='FFFFFF')
                elif avance_promedio > 0:
                    cell.fill = self.COLOR_PARCIAL
                else:
                    cell.fill = self.COLOR_PENDIENTE
                    cell.font = Font(color='FFFFFF')

            row_num += 1

        # Ajustar anchos
        sheet.column_dimensions['A'].width = 12
        for i in range(vanos_por_fila):
            sheet.column_dimensions[get_column_letter(i+2)].width = 8

    def _generar_hoja_pendientes(self, linea, torres):
        """Genera lista detallada de pendientes/condiciones."""
        from apps.campo.models import RegistroCampo

        sheet = self.workbook.create_sheet(title='Pendientes')

        # Título
        sheet['A1'] = 'LISTA DE PENDIENTES Y CONDICIONES ESPECIALES'
        sheet['A1'].font = Font(bold=True, size=12)

        # Encabezados
        headers = ['#', 'Torre', 'Tipo Pendiente', 'Descripción', 'Actividad', 'Fecha Reporte', 'Estado']
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=3, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.CELL_BORDER

        # Obtener pendientes
        registros = RegistroCampo.objects.filter(
            actividad__linea=linea,
            tiene_pendiente=True
        ).select_related(
            'actividad__torre', 'actividad__tipo_actividad'
        ).order_by('actividad__torre__numero')

        row_num = 4
        for idx, reg in enumerate(registros, start=1):
            torre = reg.actividad.torre.numero if reg.actividad.torre else '-'
            tipo = reg.get_tipo_pendiente_display() if reg.tipo_pendiente else '-'
            descripcion = reg.descripcion_pendiente or '-'
            actividad = reg.actividad.tipo_actividad.nombre
            fecha = reg.created_at.strftime('%d/%m/%Y')

            # Estado basado en si se resolvió
            if reg.actividad.porcentaje_avance >= 100:
                estado = 'Resuelto'
                estado_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            else:
                estado = 'Pendiente'
                estado_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')

            row_data = [idx, torre, tipo, descripcion, actividad, fecha, estado]
            for col_idx, value in enumerate(row_data, start=1):
                cell = sheet.cell(row=row_num, column=col_idx, value=value)
                cell.border = self.CELL_BORDER
                if col_idx == 7:  # Estado
                    cell.fill = estado_fill

            row_num += 1

        # Ajustar anchos
        widths = [5, 10, 20, 50, 25, 15, 12]
        for col_idx, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

    def _generar_hoja_resumen_con_grafico(self, linea, torres, fecha_corte):
        """Genera resumen con gráfico de torta."""
        from apps.actividades.models import Actividad

        sheet = self.workbook.create_sheet(title='Resumen')

        # Título
        sheet['A1'] = 'RESUMEN DE AVANCE'
        sheet['A1'].font = Font(bold=True, size=14)

        # Calcular estadísticas
        total_vanos = max(0, torres.count() - 1)

        actividades = Actividad.objects.filter(linea=linea)
        total_actividades = actividades.count()

        # Contar por estado
        completos = 0
        parciales = 0
        pendientes = 0

        torres_list = list(torres)
        for i in range(len(torres_list) - 1):
            torre = torres_list[i]
            acts = actividades.filter(torre=torre)
            if acts.exists():
                avance = sum(a.porcentaje_avance for a in acts) / acts.count()
                if avance >= 100:
                    completos += 1
                elif avance > 0:
                    parciales += 1
                else:
                    pendientes += 1
            else:
                pendientes += 1

        # Datos para el gráfico
        sheet['A4'] = 'Estado'
        sheet['B4'] = 'Cantidad'
        sheet['C4'] = 'Porcentaje'

        datos = [
            ('Completos (100%)', completos),
            ('Parciales (1-99%)', parciales),
            ('Pendientes (0%)', pendientes),
        ]

        for idx, (estado, cantidad) in enumerate(datos, start=5):
            sheet[f'A{idx}'] = estado
            sheet[f'B{idx}'] = cantidad
            if total_vanos > 0:
                sheet[f'C{idx}'] = f'{cantidad/total_vanos*100:.1f}%'
            else:
                sheet[f'C{idx}'] = '0%'

        # Crear gráfico de torta
        chart = PieChart()
        labels = Reference(sheet, min_col=1, min_row=5, max_row=7)
        data = Reference(sheet, min_col=2, min_row=4, max_row=7)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.title = "Distribución de Avance"

        # Mostrar porcentajes en el gráfico
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showVal = False

        sheet.add_chart(chart, "E4")

        # Resumen textual
        sheet['A12'] = 'ESTADÍSTICAS GENERALES'
        sheet['A12'].font = Font(bold=True)

        resumen = [
            ('Línea:', f'{linea.codigo} - {linea.nombre}'),
            ('Fecha de corte:', fecha_corte.strftime('%d/%m/%Y')),
            ('Total vanos:', total_vanos),
            ('Vanos completos:', f'{completos} ({completos/total_vanos*100:.1f}%)' if total_vanos else '0'),
            ('Vanos parciales:', f'{parciales} ({parciales/total_vanos*100:.1f}%)' if total_vanos else '0'),
            ('Vanos pendientes:', f'{pendientes} ({pendientes/total_vanos*100:.1f}%)' if total_vanos else '0'),
            ('', ''),
            ('AVANCE GENERAL:', f'{(completos + parciales*0.5)/total_vanos*100:.1f}%' if total_vanos else '0%'),
        ]

        for idx, (label, value) in enumerate(resumen, start=13):
            sheet[f'A{idx}'] = label
            sheet[f'A{idx}'].font = Font(bold=True) if label.startswith('AVANCE') else None
            sheet[f'B{idx}'] = value
            if label.startswith('AVANCE'):
                sheet[f'B{idx}'].font = Font(bold=True, size=16)

        # Ajustar anchos
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 12


class ReporteComparativoCuadrillas:
    """
    Genera reporte comparativo de rendimiento entre cuadrillas.
    """

    def __init__(self):
        self.workbook = None

    def generar(self, fecha_inicio, fecha_fin, linea_id=None):
        """
        Genera reporte comparativo de cuadrillas.

        Args:
            fecha_inicio: Fecha de inicio del período
            fecha_fin: Fecha de fin del período
            linea_id: UUID de línea para filtrar (opcional)

        Returns:
            BytesIO con el archivo Excel
        """
        from apps.cuadrillas.models import Cuadrilla
        from apps.actividades.models import Actividad

        self.workbook = Workbook()
        sheet = self.workbook.active
        sheet.title = 'Comparativo Cuadrillas'

        # Título
        sheet.merge_cells('A1:G1')
        sheet['A1'] = f'COMPARATIVO DE RENDIMIENTO - {fecha_inicio.strftime("%d/%m/%Y")} al {fecha_fin.strftime("%d/%m/%Y")}'
        sheet['A1'].font = Font(bold=True, size=14)
        sheet['A1'].alignment = Alignment(horizontal='center')

        # Encabezados
        headers = [
            'Cuadrilla', 'Supervisor', 'Actividades', 'Completadas',
            '% Cumplimiento', 'Avance Promedio', 'Eficiencia'
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

        # Obtener datos por cuadrilla
        cuadrillas = Cuadrilla.objects.filter(activa=True)

        row_num = 4
        for cuadrilla in cuadrillas:
            qs = Actividad.objects.filter(
                cuadrilla=cuadrilla,
                fecha_programada__gte=fecha_inicio,
                fecha_programada__lte=fecha_fin
            )

            if linea_id:
                qs = qs.filter(linea_id=linea_id)

            total = qs.count()
            if total == 0:
                continue

            completadas = qs.filter(estado='COMPLETADA').count()
            cumplimiento = (completadas / total * 100) if total > 0 else 0

            avance_promedio = sum(a.porcentaje_avance for a in qs) / total if total > 0 else 0

            # Eficiencia = avance real / días disponibles * rendimiento esperado
            dias = (fecha_fin - fecha_inicio).days + 1
            rendimiento_esperado = sum(
                a.tipo_actividad.rendimiento_estandar_vanos for a in qs
            )
            eficiencia = (avance_promedio / 100 * total / rendimiento_esperado * 100) if rendimiento_esperado > 0 else 0

            row_data = [
                cuadrilla.codigo,
                cuadrilla.supervisor.get_full_name() if cuadrilla.supervisor else '-',
                total,
                completadas,
                f'{cumplimiento:.1f}%',
                f'{avance_promedio:.1f}%',
                f'{eficiencia:.1f}%'
            ]

            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_num, column=col_idx, value=value)

            row_num += 1

        # Ajustar anchos
        widths = [15, 25, 15, 15, 15, 18, 12]
        for col_idx, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)

        return output
