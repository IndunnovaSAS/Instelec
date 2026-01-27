"""
Exporters for activity programming to Excel files.
"""
import logging
from io import BytesIO
from datetime import date, timedelta
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ProgramacionSemanalExporter:
    """
    Genera Excel formato Transelca para programación semanal:
    | Cuadrilla | Actividad | Línea | Tramo | Personal (Cédula/Tel/Cargo) | Placa |

    Formato flexible que incluye toda la información requerida.
    """

    # Estilos
    HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
    CELL_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT_ALIGNMENT = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def __init__(self):
        self.workbook = None
        self.sheet = None

    def generar_excel(self, semana_inicio, semana_fin=None, linea_id=None, cuadrilla_id=None):
        """
        Genera Excel de programación semanal.

        Args:
            semana_inicio: Date - inicio de la semana
            semana_fin: Date - fin de la semana (default: inicio + 6 días)
            linea_id: UUID - filtrar por línea (opcional)
            cuadrilla_id: UUID - filtrar por cuadrilla (opcional)

        Returns:
            BytesIO with Excel file content
        """
        from .models import Actividad
        from apps.cuadrillas.models import Cuadrilla

        if semana_fin is None:
            semana_fin = semana_inicio + timedelta(days=6)

        # Consultar actividades programadas
        qs = Actividad.objects.filter(
            fecha_programada__gte=semana_inicio,
            fecha_programada__lte=semana_fin,
            estado__in=['PENDIENTE', 'PROGRAMADA', 'EN_CURSO']
        ).select_related(
            'linea', 'torre', 'tipo_actividad', 'cuadrilla', 'tramo',
            'cuadrilla__supervisor', 'cuadrilla__vehiculo'
        ).prefetch_related(
            'cuadrilla__miembros__usuario'
        ).order_by('cuadrilla__codigo', 'fecha_programada')

        if linea_id:
            qs = qs.filter(linea_id=linea_id)

        if cuadrilla_id:
            qs = qs.filter(cuadrilla_id=cuadrilla_id)

        # Crear workbook
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = 'Programación Semanal'

        # Título
        self.sheet.merge_cells('A1:I1')
        titulo_cell = self.sheet['A1']
        titulo_cell.value = f'PROGRAMACIÓN SEMANAL - {semana_inicio.strftime("%d/%m/%Y")} al {semana_fin.strftime("%d/%m/%Y")}'
        titulo_cell.font = Font(bold=True, size=14)
        titulo_cell.alignment = self.CENTER_ALIGNMENT

        # Encabezados
        headers = [
            'Cuadrilla',
            'Fecha',
            'Actividad',
            'Aviso SAP',
            'Línea',
            'Tramo (Torre Inicio - Torre Fin)',
            'Personal',
            'Vehículo (Placa)',
            'Estado',
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = self.sheet.cell(row=3, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.CELL_BORDER
            cell.alignment = self.CENTER_ALIGNMENT

        # Datos
        row_num = 4
        cuadrilla_anterior = None

        for actividad in qs:
            # Información de cuadrilla
            cuadrilla_codigo = actividad.cuadrilla.codigo if actividad.cuadrilla else 'Sin asignar'

            # Agregar separador visual entre cuadrillas
            if cuadrilla_anterior and cuadrilla_anterior != cuadrilla_codigo:
                row_num += 1

            cuadrilla_anterior = cuadrilla_codigo

            # Información del tramo
            if actividad.tramo:
                tramo_info = f"{actividad.tramo.nombre} (T{actividad.tramo.torre_inicio.numero} - T{actividad.tramo.torre_fin.numero})"
            elif actividad.torre:
                tramo_info = f"Torre {actividad.torre.numero}"
            else:
                tramo_info = '-'

            # Información del personal
            personal_info = self._formatear_personal(actividad.cuadrilla) if actividad.cuadrilla else '-'

            # Placa del vehículo
            placa = actividad.cuadrilla.vehiculo.placa if actividad.cuadrilla and actividad.cuadrilla.vehiculo else '-'

            # Llenar fila
            row_data = [
                cuadrilla_codigo,
                actividad.fecha_programada.strftime('%d/%m/%Y'),
                actividad.tipo_actividad.nombre,
                actividad.aviso_sap or '-',
                actividad.linea.codigo,
                tramo_info,
                personal_info,
                placa,
                actividad.get_estado_display(),
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = self.sheet.cell(row=row_num, column=col_idx, value=value)
                cell.border = self.CELL_BORDER
                cell.alignment = self.LEFT_ALIGNMENT if col_idx in [3, 6, 7] else self.CENTER_ALIGNMENT

            row_num += 1

        # Ajustar anchos de columna
        column_widths = [15, 12, 30, 15, 15, 35, 50, 12, 15]
        for col_idx, width in enumerate(column_widths, start=1):
            self.sheet.column_dimensions[get_column_letter(col_idx)].width = width

        # Agregar hoja con resumen por cuadrilla
        self._agregar_hoja_resumen(qs)

        # Guardar en BytesIO
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)

        return output

    def _formatear_personal(self, cuadrilla):
        """
        Formatea la información del personal de la cuadrilla.
        Formato: Nombre (Cédula, Tel, Cargo)
        """
        if not cuadrilla:
            return '-'

        personal_lista = []
        miembros = cuadrilla.miembros.filter(activo=True).select_related('usuario')

        for miembro in miembros:
            usuario = miembro.usuario
            nombre = usuario.get_full_name()
            cedula = getattr(usuario, 'cedula', '') or ''
            telefono = getattr(usuario, 'telefono', '') or ''
            cargo = miembro.get_rol_cuadrilla_display()

            # Determinar si es CTA (Contratista de Apoyo)
            es_cta = getattr(usuario, 'es_contratista', False)
            cta_tag = ' [CTA]' if es_cta else ''

            info_personal = f"{nombre} ({cedula}, {telefono}, {cargo}){cta_tag}"
            personal_lista.append(info_personal)

        return '\n'.join(personal_lista) if personal_lista else '-'

    def _agregar_hoja_resumen(self, actividades):
        """Agrega una hoja con resumen de actividades por cuadrilla."""
        resumen_sheet = self.workbook.create_sheet(title='Resumen')

        # Título
        resumen_sheet.merge_cells('A1:D1')
        titulo_cell = resumen_sheet['A1']
        titulo_cell.value = 'RESUMEN POR CUADRILLA'
        titulo_cell.font = Font(bold=True, size=12)
        titulo_cell.alignment = self.CENTER_ALIGNMENT

        # Encabezados
        headers = ['Cuadrilla', 'Total Actividades', 'Líneas', 'Supervisor']
        for col_idx, header in enumerate(headers, start=1):
            cell = resumen_sheet.cell(row=3, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.CELL_BORDER
            cell.alignment = self.CENTER_ALIGNMENT

        # Agrupar por cuadrilla
        from collections import defaultdict
        resumen = defaultdict(lambda: {'total': 0, 'lineas': set(), 'supervisor': ''})

        for act in actividades:
            if act.cuadrilla:
                codigo = act.cuadrilla.codigo
                resumen[codigo]['total'] += 1
                resumen[codigo]['lineas'].add(act.linea.codigo)
                if act.cuadrilla.supervisor:
                    resumen[codigo]['supervisor'] = act.cuadrilla.supervisor.get_full_name()

        # Datos
        row_num = 4
        for codigo, datos in sorted(resumen.items()):
            resumen_sheet.cell(row=row_num, column=1, value=codigo).border = self.CELL_BORDER
            resumen_sheet.cell(row=row_num, column=2, value=datos['total']).border = self.CELL_BORDER
            resumen_sheet.cell(row=row_num, column=3, value=', '.join(sorted(datos['lineas']))).border = self.CELL_BORDER
            resumen_sheet.cell(row=row_num, column=4, value=datos['supervisor']).border = self.CELL_BORDER
            row_num += 1

        # Ajustar anchos
        for col_idx, width in enumerate([15, 18, 25, 30], start=1):
            resumen_sheet.column_dimensions[get_column_letter(col_idx)].width = width


class ReporteAvanceExporter:
    """
    Genera Excel con reporte de avance de servidumbre:
    - Vanos ejecutados con código de colores
    - Lista de pendientes
    - Gráfico de avance (si es posible)
    """

    # Colores para estado de vanos
    COLOR_COMPLETO = PatternFill(start_color='006400', end_color='006400', fill_type='solid')  # Verde oscuro
    COLOR_PARCIAL = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Verde claro
    COLOR_PENDIENTE = PatternFill(start_color='4169E1', end_color='4169E1', fill_type='solid')  # Azul
    COLOR_CON_PENDIENTE = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')  # Amarillo

    def __init__(self):
        self.workbook = None

    def generar_excel(self, linea_id, fecha_corte=None):
        """
        Genera reporte de avance para una línea.

        Args:
            linea_id: UUID de la línea
            fecha_corte: Date para calcular avance (default: hoy)

        Returns:
            BytesIO with Excel file content
        """
        from apps.lineas.models import Linea, Torre, Tramo
        from .models import Actividad
        from apps.campo.models import RegistroCampo

        if fecha_corte is None:
            fecha_corte = date.today()

        try:
            linea = Linea.objects.get(id=linea_id)
        except Linea.DoesNotExist:
            raise ValueError(f'Línea no encontrada: {linea_id}')

        self.workbook = Workbook()

        # Hoja principal: Estado por Torre
        self._generar_hoja_torres(linea, fecha_corte)

        # Hoja de pendientes
        self._generar_hoja_pendientes(linea)

        # Hoja de resumen
        self._generar_hoja_resumen(linea, fecha_corte)

        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)

        return output

    def _generar_hoja_torres(self, linea, fecha_corte):
        """Genera hoja con estado de torres."""
        from apps.lineas.models import Torre
        from .models import Actividad

        sheet = self.workbook.active
        sheet.title = 'Estado Torres'

        # Título
        sheet.merge_cells('A1:F1')
        titulo = sheet['A1']
        titulo.value = f'ESTADO DE AVANCE - {linea.codigo} - {linea.nombre}'
        titulo.font = Font(bold=True, size=14)

        sheet['A2'] = f'Fecha de corte: {fecha_corte.strftime("%d/%m/%Y")}'

        # Encabezados
        headers = ['Torre', 'Tipo', 'Actividades', 'Avance %', 'Estado', 'Observaciones']
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=4, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

        # Torres
        torres = Torre.objects.filter(linea=linea).order_by('numero')
        row_num = 5

        for torre in torres:
            # Obtener actividades de esta torre
            actividades = Actividad.objects.filter(
                linea=linea,
                torre=torre
            )

            total_actividades = actividades.count()
            if total_actividades > 0:
                avance_promedio = sum(a.porcentaje_avance for a in actividades) / total_actividades
            else:
                avance_promedio = 0

            # Verificar si hay pendientes
            registros_pendientes = any(
                r.tiene_pendiente
                for a in actividades
                for r in a.registros_campo.all()
            )

            # Determinar estado y color
            if avance_promedio >= 100:
                estado = 'Completo'
                fill = self.COLOR_COMPLETO
            elif avance_promedio > 0:
                estado = 'En progreso'
                fill = self.COLOR_PARCIAL
            elif registros_pendientes:
                estado = 'Con pendiente'
                fill = self.COLOR_CON_PENDIENTE
            else:
                estado = 'Pendiente'
                fill = self.COLOR_PENDIENTE

            # Observaciones
            obs_list = []
            for act in actividades:
                for reg in act.registros_campo.filter(tiene_pendiente=True):
                    obs_list.append(reg.descripcion_pendiente[:50] if reg.descripcion_pendiente else '')

            observaciones = '; '.join(obs_list) if obs_list else ''

            # Llenar fila
            sheet.cell(row=row_num, column=1, value=torre.numero)
            sheet.cell(row=row_num, column=2, value=torre.get_tipo_display())
            sheet.cell(row=row_num, column=3, value=total_actividades)
            sheet.cell(row=row_num, column=4, value=f'{avance_promedio:.1f}%')
            cell_estado = sheet.cell(row=row_num, column=5, value=estado)
            cell_estado.fill = fill
            if fill in [self.COLOR_COMPLETO, self.COLOR_PENDIENTE]:
                cell_estado.font = Font(color='FFFFFF')
            sheet.cell(row=row_num, column=6, value=observaciones)

            row_num += 1

        # Ajustar anchos
        for col_idx, width in enumerate([10, 15, 15, 12, 15, 40], start=1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

    def _generar_hoja_pendientes(self, linea):
        """Genera hoja con lista de pendientes."""
        from apps.campo.models import RegistroCampo

        sheet = self.workbook.create_sheet(title='Pendientes')

        # Título
        sheet['A1'] = 'LISTA DE PENDIENTES'
        sheet['A1'].font = Font(bold=True, size=14)

        # Encabezados
        headers = ['Torre', 'Actividad', 'Tipo Pendiente', 'Descripción', 'Fecha Reporte', 'Reportado por']
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

        # Pendientes
        registros = RegistroCampo.objects.filter(
            actividad__linea=linea,
            tiene_pendiente=True
        ).select_related(
            'actividad__torre', 'actividad__tipo_actividad', 'usuario'
        ).order_by('-created_at')

        row_num = 4
        for reg in registros:
            sheet.cell(row=row_num, column=1, value=reg.actividad.torre.numero if reg.actividad.torre else '-')
            sheet.cell(row=row_num, column=2, value=reg.actividad.tipo_actividad.nombre)
            sheet.cell(row=row_num, column=3, value=reg.get_tipo_pendiente_display())
            sheet.cell(row=row_num, column=4, value=reg.descripcion_pendiente)
            sheet.cell(row=row_num, column=5, value=reg.created_at.strftime('%d/%m/%Y'))
            sheet.cell(row=row_num, column=6, value=reg.usuario.get_full_name())
            row_num += 1

        # Ajustar anchos
        for col_idx, width in enumerate([10, 25, 20, 50, 15, 25], start=1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

    def _generar_hoja_resumen(self, linea, fecha_corte):
        """Genera hoja con resumen de avance."""
        from apps.lineas.models import Torre
        from .models import Actividad

        sheet = self.workbook.create_sheet(title='Resumen')

        # Título
        sheet['A1'] = 'RESUMEN DE AVANCE'
        sheet['A1'].font = Font(bold=True, size=14)

        # Estadísticas
        total_torres = Torre.objects.filter(linea=linea).count()

        actividades = Actividad.objects.filter(linea=linea)
        total_actividades = actividades.count()

        completadas = actividades.filter(estado='COMPLETADA').count()
        en_curso = actividades.filter(estado='EN_CURSO').count()
        pendientes = actividades.filter(estado__in=['PENDIENTE', 'PROGRAMADA']).count()

        avance_general = (completadas / total_actividades * 100) if total_actividades > 0 else 0

        # Mostrar estadísticas
        datos = [
            ('Línea:', f'{linea.codigo} - {linea.nombre}'),
            ('Fecha de corte:', fecha_corte.strftime('%d/%m/%Y')),
            ('', ''),
            ('Total Torres:', total_torres),
            ('Total Actividades:', total_actividades),
            ('', ''),
            ('Completadas:', f'{completadas} ({completadas/total_actividades*100:.1f}%)' if total_actividades else '0'),
            ('En Curso:', f'{en_curso} ({en_curso/total_actividades*100:.1f}%)' if total_actividades else '0'),
            ('Pendientes:', f'{pendientes} ({pendientes/total_actividades*100:.1f}%)' if total_actividades else '0'),
            ('', ''),
            ('AVANCE GENERAL:', f'{avance_general:.1f}%'),
        ]

        row_num = 3
        for label, value in datos:
            sheet.cell(row=row_num, column=1, value=label).font = Font(bold=True)
            sheet.cell(row=row_num, column=2, value=value)
            row_num += 1

        # Destacar avance general
        sheet.cell(row=row_num-1, column=2).font = Font(bold=True, size=16)

        # Ajustar anchos
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 40
