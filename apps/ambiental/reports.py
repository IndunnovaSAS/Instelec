"""Generador de informes ambientales mensuales."""

import io
from datetime import date
from decimal import Decimal
from typing import Any

from django.template.loader import render_to_string
from weasyprint import HTML, CSS
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from apps.ambiental.models import InformeAmbiental, PermisoServidumbre
from apps.actividades.models import Actividad
from apps.campo.models import RegistroCampo, Evidencia
from apps.lineas.models import Linea


class InformeAmbientalGenerator:
    """Genera informes ambientales en PDF y Excel."""

    def __init__(self, informe: InformeAmbiental):
        self.informe = informe
        self.linea = informe.linea
        self.anio = informe.periodo_anio
        self.mes = informe.periodo_mes
        self.data = {}

    def consolidar_datos(self) -> dict[str, Any]:
        """Consolida todos los datos para el informe."""
        # Obtener actividades del período
        actividades = Actividad.objects.filter(
            linea=self.linea,
            fecha_programada__year=self.anio,
            fecha_programada__month=self.mes,
            estado='COMPLETADA'
        ).select_related('tipo_actividad', 'torre', 'cuadrilla')

        # Obtener registros de campo
        registros = RegistroCampo.objects.filter(
            actividad__in=actividades,
            sincronizado=True
        ).select_related('actividad', 'usuario')

        # Calcular métricas
        self.data = {
            'informe': self.informe,
            'linea': self.linea,
            'periodo': f"{self._get_mes_nombre(self.mes)} {self.anio}",
            'fecha_generacion': date.today(),

            # Resumen de actividades
            'actividades': {
                'total': actividades.count(),
                'por_tipo': self._agrupar_por_tipo(actividades),
                'por_cuadrilla': self._agrupar_por_cuadrilla(actividades),
            },

            # Intervención ambiental
            'ambiental': {
                'area_intervenida': self._calcular_area_intervenida(registros),
                'vegetacion_podada': self._calcular_vegetacion(registros),
                'residuos_generados': self._calcular_residuos(registros),
                'disposicion_residuos': self._obtener_disposicion_residuos(),
            },

            # Permisos y servidumbres
            'permisos': {
                'vigentes': self._obtener_permisos_vigentes(),
                'por_vencer': self._obtener_permisos_por_vencer(),
                'vencidos': self._obtener_permisos_vencidos(),
            },

            # Incidentes ambientales
            'incidentes': self._obtener_incidentes(),

            # Evidencias fotográficas
            'evidencias': self._obtener_evidencias_resumen(registros),

            # Torres intervenidas
            'torres': self._obtener_torres_intervenidas(actividades),
        }

        return self.data

    def render_html(self) -> str:
        """Renderiza el informe como HTML."""
        if not self.data:
            self.consolidar_datos()

        return render_to_string('ambiental/reports/informe_mensual.html', self.data)

    def generar_pdf(self) -> bytes:
        """Genera el informe en formato PDF."""
        html_content = self.render_html()

        # CSS para el PDF
        css = CSS(string='''
            @page {
                size: letter;
                margin: 2cm;
                @top-center { content: "Informe Ambiental Mensual"; }
                @bottom-right { content: "Página " counter(page) " de " counter(pages); }
            }
            body { font-family: Arial, sans-serif; font-size: 10pt; }
            h1 { color: #1e40af; border-bottom: 2px solid #1e40af; padding-bottom: 10px; }
            h2 { color: #374151; margin-top: 20px; }
            table { width: 100%; border-collapse: collapse; margin: 10px 0; }
            th, td { border: 1px solid #d1d5db; padding: 8px; text-align: left; }
            th { background-color: #f3f4f6; font-weight: bold; }
            .metric { background-color: #eff6ff; padding: 15px; border-radius: 8px; margin: 10px 0; }
            .metric-value { font-size: 24pt; font-weight: bold; color: #1e40af; }
            .alert { background-color: #fef2f2; border-left: 4px solid #ef4444; padding: 10px; }
            .success { background-color: #f0fdf4; border-left: 4px solid #22c55e; padding: 10px; }
        ''')

        html = HTML(string=html_content)
        return html.write_pdf(stylesheets=[css])

    def generar_excel(self) -> bytes:
        """Genera el informe en formato Excel."""
        if not self.data:
            self.consolidar_datos()

        wb = Workbook()

        # Hoja de resumen
        self._crear_hoja_resumen(wb)

        # Hoja de actividades
        self._crear_hoja_actividades(wb)

        # Hoja de torres
        self._crear_hoja_torres(wb)

        # Hoja de permisos
        self._crear_hoja_permisos(wb)

        # Guardar en buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer.getvalue()

    def _crear_hoja_resumen(self, wb: Workbook):
        """Crea la hoja de resumen ejecutivo."""
        ws = wb.active
        ws.title = "Resumen"

        # Estilos
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        title_font = Font(bold=True, size=12)

        # Título
        ws.merge_cells('A1:F1')
        ws['A1'] = f"INFORME AMBIENTAL - {self.linea.codigo}"
        ws['A1'].font = Font(bold=True, size=16)

        ws.merge_cells('A2:F2')
        ws['A2'] = f"Período: {self.data['periodo']}"

        # Métricas principales
        row = 4
        metrics = [
            ("Total Actividades", self.data['actividades']['total']),
            ("Área Intervenida (ha)", self.data['ambiental']['area_intervenida']),
            ("Vegetación Podada (m³)", self.data['ambiental']['vegetacion_podada']),
            ("Residuos Generados (kg)", self.data['ambiental']['residuos_generados']),
            ("Permisos Vigentes", len(self.data['permisos']['vigentes'])),
            ("Incidentes", len(self.data['incidentes'])),
        ]

        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = title_font
            ws[f'B{row}'] = value
            row += 1

        # Ajustar anchos
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

    def _crear_hoja_actividades(self, wb: Workbook):
        """Crea la hoja de detalle de actividades."""
        ws = wb.create_sheet("Actividades")

        headers = ["Fecha", "Torre", "Tipo Actividad", "Cuadrilla", "Estado", "Observaciones"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")

        # Datos (simplificado - en producción usar datos reales)
        row = 2
        for tipo, datos in self.data['actividades']['por_tipo'].items():
            ws.cell(row=row, column=1, value=str(date.today()))
            ws.cell(row=row, column=2, value="-")
            ws.cell(row=row, column=3, value=tipo)
            ws.cell(row=row, column=4, value="-")
            ws.cell(row=row, column=5, value="Completada")
            ws.cell(row=row, column=6, value=f"{datos['cantidad']} actividades")
            row += 1

    def _crear_hoja_torres(self, wb: Workbook):
        """Crea la hoja de torres intervenidas."""
        ws = wb.create_sheet("Torres")

        headers = ["Torre", "Tipo", "Latitud", "Longitud", "Actividades", "Última Intervención"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")

    def _crear_hoja_permisos(self, wb: Workbook):
        """Crea la hoja de permisos de servidumbre."""
        ws = wb.create_sheet("Permisos")

        headers = ["Propietario", "Predio", "Fecha Inicio", "Fecha Vencimiento", "Estado", "Área (ha)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")

    # Métodos auxiliares de cálculo
    def _get_mes_nombre(self, mes: int) -> str:
        meses = [
            '', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
            'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
        ]
        return meses[mes]

    def _agrupar_por_tipo(self, actividades) -> dict:
        resultado = {}
        for act in actividades:
            tipo = act.tipo_actividad.nombre
            if tipo not in resultado:
                resultado[tipo] = {'cantidad': 0, 'categoria': act.tipo_actividad.categoria}
            resultado[tipo]['cantidad'] += 1
        return resultado

    def _agrupar_por_cuadrilla(self, actividades) -> dict:
        resultado = {}
        for act in actividades:
            if act.cuadrilla:
                nombre = act.cuadrilla.nombre
                if nombre not in resultado:
                    resultado[nombre] = 0
                resultado[nombre] += 1
        return resultado

    def _calcular_area_intervenida(self, registros) -> Decimal:
        total = Decimal('0')
        for reg in registros:
            if 'area_intervenida' in reg.datos_formulario:
                try:
                    total += Decimal(str(reg.datos_formulario['area_intervenida']))
                except (ValueError, TypeError):
                    pass
        return total

    def _calcular_vegetacion(self, registros) -> Decimal:
        total = Decimal('0')
        for reg in registros:
            if 'volumen_vegetacion' in reg.datos_formulario:
                try:
                    total += Decimal(str(reg.datos_formulario['volumen_vegetacion']))
                except (ValueError, TypeError):
                    pass
        return total

    def _calcular_residuos(self, registros) -> Decimal:
        total = Decimal('0')
        for reg in registros:
            if 'residuos_kg' in reg.datos_formulario:
                try:
                    total += Decimal(str(reg.datos_formulario['residuos_kg']))
                except (ValueError, TypeError):
                    pass
        return total

    def _obtener_disposicion_residuos(self) -> list:
        return [
            {'tipo': 'Vegetación', 'cantidad': 0, 'disposicion': 'Compostaje'},
            {'tipo': 'Metálicos', 'cantidad': 0, 'disposicion': 'Reciclaje'},
            {'tipo': 'Peligrosos', 'cantidad': 0, 'disposicion': 'Gestor autorizado'},
        ]

    def _obtener_permisos_vigentes(self) -> list:
        return list(PermisoServidumbre.objects.filter(
            linea=self.linea,
            fecha_vencimiento__gte=date.today(),
            estado='VIGENTE'
        ))

    def _obtener_permisos_por_vencer(self) -> list:
        from datetime import timedelta
        fecha_limite = date.today() + timedelta(days=30)
        return list(PermisoServidumbre.objects.filter(
            linea=self.linea,
            fecha_vencimiento__lte=fecha_limite,
            fecha_vencimiento__gte=date.today(),
            estado='VIGENTE'
        ))

    def _obtener_permisos_vencidos(self) -> list:
        return list(PermisoServidumbre.objects.filter(
            linea=self.linea,
            fecha_vencimiento__lt=date.today()
        ).exclude(estado='RENOVADO'))

    def _obtener_incidentes(self) -> list:
        # En producción, obtener de un modelo de incidentes
        return []

    def _obtener_evidencias_resumen(self, registros) -> dict:
        total = Evidencia.objects.filter(registro_campo__in=registros).count()
        por_tipo = {
            'ANTES': Evidencia.objects.filter(registro_campo__in=registros, tipo='ANTES').count(),
            'DURANTE': Evidencia.objects.filter(registro_campo__in=registros, tipo='DURANTE').count(),
            'DESPUES': Evidencia.objects.filter(registro_campo__in=registros, tipo='DESPUES').count(),
        }
        return {'total': total, 'por_tipo': por_tipo}

    def _obtener_torres_intervenidas(self, actividades) -> list:
        torres = set()
        for act in actividades:
            torres.add(act.torre)
        return list(torres)
