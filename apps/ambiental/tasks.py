"""
Celery tasks for environmental report generation.
"""
from celery import shared_task
from celery.utils.log import get_task_logger
from datetime import date, timedelta

logger = get_task_logger(__name__)


@shared_task
def generar_informe_ambiental(informe_id: str):
    """
    Generate environmental report PDF and Excel.
    """
    from apps.ambiental.models import InformeAmbiental
    from apps.actividades.models import Actividad
    from apps.campo.models import RegistroCampo
    from apps.core.utils import upload_to_gcs
    from django.utils import timezone

    logger.info(f"Generating environmental report {informe_id}")

    informe = InformeAmbiental.objects.select_related('linea').get(id=informe_id)

    # Get completed activities for the period
    actividades = Actividad.objects.filter(
        linea=informe.linea,
        fecha_programada__year=informe.periodo_anio,
        fecha_programada__month=informe.periodo_mes,
        estado='COMPLETADA'
    ).select_related('torre', 'tipo_actividad')

    # Get field records
    registros = RegistroCampo.objects.filter(
        actividad__in=actividades,
        sincronizado=True
    ).prefetch_related('evidencias')

    # Update summary
    informe.total_actividades = actividades.count()
    informe.total_podas = actividades.filter(
        tipo_actividad__categoria='PODA'
    ).count()

    # Generate PDF
    pdf_content = generar_pdf_informe(informe, actividades, registros)
    pdf_path = f"informes/ambiental/{informe.periodo_anio}/{informe.periodo_mes}/{informe.linea.codigo}.pdf"
    pdf_url = upload_to_gcs(pdf_content, pdf_path)

    # Generate Excel
    excel_content = generar_excel_informe(informe, actividades, registros)
    excel_path = f"informes/ambiental/{informe.periodo_anio}/{informe.periodo_mes}/{informe.linea.codigo}.xlsx"
    excel_url = upload_to_gcs(excel_content, excel_path)

    # Update informe
    informe.url_pdf = pdf_url
    informe.url_excel = excel_url
    informe.estado = InformeAmbiental.Estado.EN_REVISION
    informe.save()

    logger.info(f"Environmental report {informe_id} generated successfully")
    return {'pdf_url': pdf_url, 'excel_url': excel_url}


def generar_pdf_informe(informe, actividades, registros):
    """Generate PDF report using WeasyPrint."""
    from django.template.loader import render_to_string
    from weasyprint import HTML
    import io

    # Render HTML template
    html_content = render_to_string('ambiental/pdf/informe.html', {
        'informe': informe,
        'actividades': actividades,
        'registros': registros,
    })

    # Convert to PDF
    pdf_buffer = io.BytesIO()
    HTML(string=html_content).write_pdf(pdf_buffer)
    pdf_buffer.seek(0)

    return pdf_buffer.getvalue()


def generar_excel_informe(informe, actividades, registros):
    """Generate Excel report using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    # Header
    ws['A1'] = f"Informe Ambiental - {informe.linea.codigo}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Período: {informe.periodo_mes}/{informe.periodo_anio}"

    # Summary
    ws['A4'] = "RESUMEN"
    ws['A4'].font = Font(bold=True)
    ws['A5'] = "Total Actividades"
    ws['B5'] = informe.total_actividades
    ws['A6'] = "Total Podas"
    ws['B6'] = informe.total_podas

    # Activities sheet
    ws_act = wb.create_sheet("Actividades")
    headers = ['Torre', 'Tipo', 'Fecha', 'Estado', 'Cuadrilla']
    for col, header in enumerate(headers, 1):
        cell = ws_act.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    for row, act in enumerate(actividades, 2):
        ws_act.cell(row=row, column=1, value=act.torre.numero)
        ws_act.cell(row=row, column=2, value=act.tipo_actividad.nombre)
        ws_act.cell(row=row, column=3, value=act.fecha_programada.isoformat())
        ws_act.cell(row=row, column=4, value=act.estado)
        ws_act.cell(row=row, column=5, value=act.cuadrilla.codigo if act.cuadrilla else '')

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer.getvalue()


@shared_task
def verificar_permisos_vencidos():
    """
    Check for expired or soon-to-expire easement permits.
    Runs daily to alert about permits requiring attention.
    """
    from .models import PermisoServidumbre

    hoy = date.today()
    fecha_limite = hoy + timedelta(days=30)

    # Permits expiring in 30 days
    por_vencer = PermisoServidumbre.objects.filter(
        fecha_vencimiento__lte=fecha_limite,
        fecha_vencimiento__gte=hoy,
        estado='VIGENTE'
    ).select_related('linea')

    # Already expired
    vencidos = PermisoServidumbre.objects.filter(
        fecha_vencimiento__lt=hoy,
        estado='VIGENTE'
    ).select_related('linea')

    # Update status of expired permits
    for permiso in vencidos:
        permiso.estado = 'VENCIDO'
        permiso.save(update_fields=['estado'])

    alertas = {
        'por_vencer': [
            {
                'permiso_id': str(p.id),
                'linea': p.linea.codigo,
                'propietario': p.nombre_propietario,
                'dias_restantes': (p.fecha_vencimiento - hoy).days
            }
            for p in por_vencer
        ],
        'vencidos': [str(p.id) for p in vencidos]
    }

    if por_vencer.exists() or vencidos.exists():
        logger.warning(f"Permit alerts: {len(alertas['por_vencer'])} expiring, {len(alertas['vencidos'])} expired")

    return alertas


@shared_task
def generar_informes_periodo(anio: int, mes: int):
    """
    Generate environmental reports for all active lines.
    Runs on the 1st of each month.
    """
    from apps.lineas.models import Linea
    from .models import InformeAmbiental

    lineas = Linea.objects.filter(activa=True)
    generados = []

    for linea in lineas:
        informe, created = InformeAmbiental.objects.get_or_create(
            linea=linea,
            periodo_anio=anio,
            periodo_mes=mes,
            defaults={'estado': 'PENDIENTE'}
        )

        if created:
            generar_informe_ambiental.delay(str(informe.id))
            generados.append(linea.codigo)

    logger.info(f"Queued {len(generados)} environmental reports for {mes}/{anio}")
    return generados
